from __future__ import annotations

import base64
import hashlib
import hmac
import json
import os
import re
import sqlite3
import shutil
import struct
import uuid
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path

import yaml
from fastapi import (
    BackgroundTasks,
    FastAPI,
    File,
    Form,
    HTTPException,
    Request,
    Query,
    UploadFile,
)
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from starlette.responses import RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

from statistic_harness.core import evaluation as eval_core
from statistic_harness.core.evaluation import evaluate_report
from statistic_harness.core.known_issue_compiler import compile_known_issues
from statistic_harness.core.auth import (
    AuthUser,
    generate_api_key,
    generate_session_token,
    hash_password,
    hash_token,
    is_api_key,
    normalize_email,
    verify_password,
)
from statistic_harness.core.pipeline import Pipeline
from statistic_harness.core.report import build_report, write_report
from statistic_harness.core.tenancy import get_tenant_context
from statistic_harness.core.utils import (
    DEFAULT_TENANT_ID,
    auth_enabled,
    file_sha256,
    atomic_write_text,
    json_dumps,
    max_upload_bytes,
    now_iso,
    safe_join,
    scope_key,
    stable_hash,
    vector_store_enabled,
)
from statistic_harness.core.upload_cas import (
    blob_path as upload_blob_path,
    promote_quarantine_file,
    quarantine_dir as upload_quarantine_dir,
)
from statistic_harness.core.vector_store import VectorStore, hash_embedding

app = FastAPI()
BASE_DIR = Path(__file__).resolve().parent
TEMPLATES = Jinja2Templates(directory=str(BASE_DIR / "templates"))
app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")

_DEFAULT_CSP = (
    "default-src 'self'; "
    "base-uri 'self'; "
    "object-src 'none'; "
    "frame-ancestors 'none'; "
    "img-src 'self' data:; "
    "style-src 'self' 'unsafe-inline'; "
    "script-src 'self' 'unsafe-inline'; "
    "connect-src 'self'; "
    "form-action 'self'"
)


def _apply_security_headers(headers) -> None:
    # `headers` is a Starlette MutableHeaders at runtime; keep this duck-typed for unit testing.
    def ensure(key: str, value: str) -> None:
        if headers.get(key) is None:
            headers[key] = value

    ensure("X-Content-Type-Options", "nosniff")
    ensure("Referrer-Policy", "no-referrer")
    ensure("X-Frame-Options", "DENY")
    ensure("Content-Security-Policy", _DEFAULT_CSP)
    ensure("Cross-Origin-Opener-Policy", "same-origin")
    ensure("Cross-Origin-Resource-Policy", "same-origin")
    ensure("Permissions-Policy", "geolocation=(), microphone=(), camera=()")


TENANT_CTX = get_tenant_context()
APPDATA_DIR = TENANT_CTX.tenant_root
pipeline = Pipeline(TENANT_CTX.appdata_root, Path("plugins"), tenant_id=TENANT_CTX.tenant_id)
KNOWN_ISSUES_DIR = APPDATA_DIR / "known_issues"
AUTH_ENABLED = auth_enabled()
SESSION_COOKIE_NAME = "stat_harness_session"


def _session_ttl_hours() -> int:
    raw = os.environ.get("STAT_HARNESS_SESSION_TTL_HOURS", "").strip()
    if not raw:
        return 24
    try:
        value = int(raw)
    except ValueError:
        return 24
    return max(1, value)


def _expires_at(hours: int) -> str:
    return (datetime.now(timezone.utc) + timedelta(hours=hours)).isoformat()


def _is_expired(expires_at: str | None) -> bool:
    if not expires_at:
        return False
    try:
        return datetime.fromisoformat(expires_at) <= datetime.now(timezone.utc)
    except ValueError:
        return False


@dataclass
class AuthResult:
    user: AuthUser
    session_id: int | None = None
    rotated_token: str | None = None
    api_key_id: int | None = None


def _extract_token(request: Request) -> tuple[str | None, str | None]:
    auth_header = request.headers.get("Authorization", "").strip()
    if auth_header.lower().startswith("bearer "):
        token = auth_header.split(" ", 1)[1].strip()
        return token, "header"
    api_key = request.headers.get("X-API-Key", "").strip()
    if api_key:
        return api_key, "api_key"
    cookie_token = request.cookies.get(SESSION_COOKIE_NAME, "").strip()
    if cookie_token:
        return cookie_token, "cookie"
    return None, None


def _current_tenant_id() -> str:
    return TENANT_CTX.tenant_id


def _authenticate_request(request: Request) -> AuthResult | None:
    token, source = _extract_token(request)
    if not token:
        return None
    now = now_iso()
    tenant_id = _current_tenant_id()
    if source == "api_key" or is_api_key(token):
        key_hash = hash_token(token)
        row = pipeline.storage.fetch_api_key_by_hash(key_hash, tenant_id=tenant_id)
        if not row or row.get("revoked_at") or row.get("disabled_at"):
            return None
        membership = pipeline.storage.fetch_membership(int(row["user_id"]), tenant_id)
        if not membership:
            return None
        pipeline.storage.touch_api_key(int(row["key_id"]), now)
        user = AuthUser(
            user_id=int(row["user_id"]),
            email=str(row["email"]),
            name=row.get("user_name"),
            is_admin=bool(row.get("is_admin")),
            tenant_id=tenant_id,
        )
        return AuthResult(user=user, api_key_id=int(row["key_id"]))

    token_hash = hash_token(token)
    session = pipeline.storage.fetch_session_by_hash(token_hash, tenant_id=tenant_id)
    if not session or session.get("revoked_at") or session.get("disabled_at"):
        return None
    if _is_expired(session.get("expires_at")):
        return None
    membership = pipeline.storage.fetch_membership(int(session["user_id"]), tenant_id)
    if not membership:
        return None

    rotated_token = None
    if source == "cookie":
        rotated_token = generate_session_token()
        pipeline.storage.rotate_session(
            int(session["session_id"]),
            hash_token(rotated_token),
            now,
            _expires_at(_session_ttl_hours()),
        )
    else:
        pipeline.storage.touch_session(int(session["session_id"]), now)
    user = AuthUser(
        user_id=int(session["user_id"]),
        email=str(session["email"]),
        name=session.get("name"),
        is_admin=bool(session.get("is_admin")),
        tenant_id=tenant_id,
    )
    return AuthResult(user=user, session_id=int(session["session_id"]), rotated_token=rotated_token)


def _auth_required_response(request: Request) -> JSONResponse | RedirectResponse:
    if request.url.path.startswith("/api/"):
        return JSONResponse(status_code=401, content={"error": "unauthorized"})
    params = f"next={request.url.path}"
    return RedirectResponse(url=f"/login?{params}", status_code=303)


def _require_admin(request: Request) -> AuthUser:
    user = getattr(request.state, "user", None)
    if not user:
        raise HTTPException(status_code=401, detail="unauthorized")
    if user.is_admin:
        return user
    membership = pipeline.storage.fetch_membership(int(user.user_id))
    if membership and membership.get("role") == "admin":
        return user
    raise HTTPException(status_code=403, detail="admin required")


@app.middleware("http")
async def security_headers_middleware(request: Request, call_next):
    response = await call_next(request)
    _apply_security_headers(response.headers)
    return response


@app.middleware("http")
async def auth_middleware(request: Request, call_next):
    if not AUTH_ENABLED:
        return await call_next(request)
    path = request.url.path
    if path.startswith("/static"):
        return await call_next(request)
    if path in {"/login", "/logout", "/bootstrap"}:
        return await call_next(request)
    if pipeline.storage.count_users() == 0 and path != "/bootstrap":
        return RedirectResponse(url="/bootstrap", status_code=303)
    auth = _authenticate_request(request)
    if not auth:
        return _auth_required_response(request)
    request.state.user = auth.user
    request.state.session_token = auth.rotated_token
    response = await call_next(request)
    if auth.rotated_token:
        response.set_cookie(
            SESSION_COOKIE_NAME,
            auth.rotated_token,
            httponly=True,
            samesite="lax",
        )
    return response
ROLE_KEYS = [
    "queue_time",
    "start_time",
    "end_time",
    "process_id",
    "process_name",
    "module_code",
    "user_id",
    "dependency_id",
    "master_id",
    "host_id",
    "status",
]


def _parse_filter_params(
    project_ids: str = "",
    dataset_ids: str = "",
    dataset_version_ids: str = "",
    raw_format_ids: str = "",
    created_after: str = "",
    created_before: str = "",
) -> dict[str, object]:
    def parse_csv(value: str) -> list[str]:
        return [item.strip() for item in value.split(",") if item.strip()]

    filters: dict[str, object] = {}
    proj_list = sorted(parse_csv(project_ids))
    if proj_list:
        filters["project_ids"] = proj_list
    ds_list = sorted(parse_csv(dataset_ids))
    if ds_list:
        filters["dataset_ids"] = ds_list
    dv_list = sorted(parse_csv(dataset_version_ids))
    if dv_list:
        filters["dataset_version_ids"] = dv_list
    rf_list = sorted({int(v) for v in parse_csv(raw_format_ids) if v.isdigit()})
    if rf_list:
        filters["raw_format_ids"] = rf_list
    if created_after:
        filters["created_after"] = created_after
    if created_before:
        filters["created_before"] = created_before
    return filters


def _ground_truth_template(report: dict) -> str:
    features = [
        f.get("feature")
        for f in report.get("plugins", {})
        .get("analysis_gaussian_knockoffs", {})
        .get("findings", [])
    ]
    template = {
        "strict": False,
        "features": [f for f in features if f],
        "changepoints": [],
        "dependence_shift_pairs": [],
        "anomalies": [],
        "min_anomaly_hits": 0,
        "changepoint_tolerance": 3,
    }
    return yaml.safe_dump(template)


def _known_issues_paths(sha256: str) -> tuple[Path, Path]:
    KNOWN_ISSUES_DIR.mkdir(parents=True, exist_ok=True)
    json_path = KNOWN_ISSUES_DIR / f"{sha256}.json"
    yaml_path = KNOWN_ISSUES_DIR / f"{sha256}.yaml"
    return json_path, yaml_path


def _vector_store() -> tuple[VectorStore | None, str | None]:
    if not vector_store_enabled():
        return None, "Vector store is disabled (set STAT_HARNESS_ENABLE_VECTOR_STORE=1)."
    try:
        store = VectorStore(TENANT_CTX.db_path, tenant_id=TENANT_CTX.tenant_id)
        return store, None
    except RuntimeError as exc:
        return None, str(exc)


def _cursor_secret() -> bytes:
    raw = os.environ.get("STAT_HARNESS_VECTOR_CURSOR_SECRET", "").strip()
    if raw:
        return raw.encode("utf-8")
    fallback = f"{TENANT_CTX.db_path}:{TENANT_CTX.tenant_id}".encode("utf-8")
    return hashlib.sha256(fallback).digest()


def _b64encode(data: bytes) -> str:
    return base64.urlsafe_b64encode(data).decode("ascii").rstrip("=")


def _b64decode(token: str) -> bytes:
    padding = "=" * (-len(token) % 4)
    return base64.urlsafe_b64decode(token + padding)


def _encode_vector(values: list[float]) -> str:
    packed = struct.pack(f"{len(values)}f", *values)
    return _b64encode(packed)


def _decode_vector(token: str, dimensions: int | None = None) -> list[float]:
    raw = _b64decode(token)
    if len(raw) % 4 != 0:
        raise ValueError("Invalid vector encoding")
    count = len(raw) // 4
    if dimensions is not None and count != int(dimensions):
        raise ValueError("Vector dimensions mismatch")
    return list(struct.unpack(f"{count}f", raw))


def _encode_cursor(payload: dict[str, Any]) -> str:
    data = json.dumps(payload, separators=(",", ":"), sort_keys=True).encode("utf-8")
    sig = hmac.new(_cursor_secret(), data, hashlib.sha256).digest()
    return f"{_b64encode(data)}.{_b64encode(sig)}"


def _decode_cursor(token: str) -> dict[str, Any]:
    try:
        data_b64, sig_b64 = token.split(".", 1)
    except ValueError as exc:
        raise ValueError("Invalid cursor token") from exc
    data = _b64decode(data_b64)
    sig = _b64decode(sig_b64)
    expected = hmac.new(_cursor_secret(), data, hashlib.sha256).digest()
    if not hmac.compare_digest(sig, expected):
        raise ValueError("Invalid cursor signature")
    payload = json.loads(data.decode("utf-8"))
    if payload.get("v") != 1:
        raise ValueError("Unsupported cursor version")
    if payload.get("tenant_id") != TENANT_CTX.tenant_id:
        raise ValueError("Cursor tenant mismatch")
    return payload


def _vector_default_collection(
    collection: str, collections: list[dict[str, Any]]
) -> str:
    if collection:
        return collection
    if collections:
        return str(collections[0].get("name") or "")
    return ""


def _parse_vector(text: str) -> list[float]:
    parts = [part for part in text.replace("\n", ",").split(",") if part.strip()]
    if not parts:
        raise ValueError("Vector is empty")
    values: list[float] = []
    for part in parts:
        values.append(float(part.strip()))
    return values


def _resolve_vector_dimensions(store: VectorStore, collection: str, dimensions: int | None) -> int:
    if dimensions is not None:
        return int(dimensions)
    dims = store.collection_dimensions(collection)
    if not dims:
        raise ValueError("Unknown collection or dimensions required")
    if len(dims) > 1:
        raise ValueError("Multiple dimensions found; provide dimensions")
    return int(dims[0])


def _vector_k_max() -> int:
    raw = os.environ.get("STAT_HARNESS_VECTOR_K_MAX", "").strip()
    if not raw:
        return 1000
    try:
        value = int(raw)
    except ValueError:
        return 1000
    return max(1, value)


def _normalize_k(value: int | str | None) -> int:
    try:
        k = int(value or 10)
    except (TypeError, ValueError):
        k = 10
    if k < 1:
        k = 1
    max_k = _vector_k_max()
    if k > max_k:
        k = max_k
    return k


def _normalize_offset(value: int | str | None) -> int:
    try:
        offset = int(value or 0)
    except (TypeError, ValueError):
        offset = 0
    if offset < 0:
        offset = 0
    return offset


@app.get("/login", response_class=HTMLResponse)
async def login_view(request: Request, next: str = "/") -> HTMLResponse:
    if not AUTH_ENABLED:
        return RedirectResponse(url=next or "/", status_code=303)
    bootstrap = pipeline.storage.count_users() == 0
    return TEMPLATES.TemplateResponse(
        "login.html",
        {"request": request, "next": next, "error": None, "bootstrap": bootstrap},
    )


@app.post("/login")
async def login_action(
    request: Request,
    email: str = Form(...),
    password: str = Form(...),
    next: str = Form("/"),
) -> HTMLResponse:
    if not AUTH_ENABLED:
        return RedirectResponse(url=next or "/", status_code=303)
    user = pipeline.storage.fetch_user_by_email(normalize_email(email))
    if not user or user.get("disabled_at"):
        return TEMPLATES.TemplateResponse(
            "login.html",
            {
                "request": request,
                "next": next,
                "error": "Invalid credentials",
                "bootstrap": pipeline.storage.count_users() == 0,
            },
            status_code=401,
        )
    membership = pipeline.storage.fetch_membership(int(user["user_id"]))
    if not membership:
        return TEMPLATES.TemplateResponse(
            "login.html",
            {
                "request": request,
                "next": next,
                "error": "Access denied",
                "bootstrap": pipeline.storage.count_users() == 0,
            },
            status_code=403,
        )
    if not verify_password(password, user.get("password_hash") or ""):
        return TEMPLATES.TemplateResponse(
            "login.html",
            {
                "request": request,
                "next": next,
                "error": "Invalid credentials",
                "bootstrap": pipeline.storage.count_users() == 0,
            },
            status_code=401,
        )
    token = generate_session_token()
    pipeline.storage.create_session(
        int(user["user_id"]),
        hash_token(token),
        now_iso(),
        _expires_at(_session_ttl_hours()),
    )
    response = RedirectResponse(url=next or "/", status_code=303)
    response.set_cookie(
        SESSION_COOKIE_NAME,
        token,
        httponly=True,
        samesite="lax",
    )
    return response


@app.get("/logout")
async def logout_view(request: Request) -> RedirectResponse:
    if AUTH_ENABLED:
        token = request.cookies.get(SESSION_COOKIE_NAME, "").strip()
        if token:
            pipeline.storage.revoke_session(hash_token(token))
    response = RedirectResponse(url="/login", status_code=303)
    response.delete_cookie(SESSION_COOKIE_NAME)
    return response


@app.get("/bootstrap", response_class=HTMLResponse)
async def bootstrap_view(request: Request) -> HTMLResponse:
    if not AUTH_ENABLED:
        return RedirectResponse(url="/", status_code=303)
    if pipeline.storage.count_users() > 0:
        return RedirectResponse(url="/login", status_code=303)
    return TEMPLATES.TemplateResponse(
        "bootstrap.html",
        {"request": request, "error": None},
    )


@app.post("/bootstrap")
async def bootstrap_action(
    request: Request,
    email: str = Form(...),
    password: str = Form(...),
    name: str = Form(""),
) -> HTMLResponse:
    if not AUTH_ENABLED:
        return RedirectResponse(url="/", status_code=303)
    if pipeline.storage.count_users() > 0:
        return RedirectResponse(url="/login", status_code=303)
    normalized = normalize_email(email)
    user_id = pipeline.storage.create_user(
        normalized, hash_password(password), name or None, True, now_iso()
    )
    pipeline.storage.ensure_membership(user_id, "admin", now_iso())
    token = generate_session_token()
    pipeline.storage.create_session(
        user_id,
        hash_token(token),
        now_iso(),
        _expires_at(_session_ttl_hours()),
    )
    response = RedirectResponse(url="/", status_code=303)
    response.set_cookie(
        SESSION_COOKIE_NAME,
        token,
        httponly=True,
        samesite="lax",
    )
    return response


@app.get("/vectors", response_class=HTMLResponse)
async def vectors_view(
    request: Request,
    collection: str = "",
    text: str = "",
    k: int = 10,
    dimensions: int | None = None,
) -> HTMLResponse:
    store, error = _vector_store()
    collections = store.list_collections() if store else []
    default_collection = _vector_default_collection(collection, collections)
    k = _normalize_k(k)
    return TEMPLATES.TemplateResponse(
        "vectors.html",
        {
            "request": request,
            "error": error,
            "collections": collections,
            "results": None,
            "query": {
                "collection": collection,
                "text": text,
                "k": k,
                "offset": 0,
                "dimensions": dimensions,
            },
            "default_collection": default_collection,
            "pagination": None,
        },
    )


@app.post("/vectors/query", response_class=HTMLResponse)
async def vectors_query(
    request: Request,
    collection: str = Form(...),
    text: str = Form(""),
    vector: str = Form(""),
    k: int = Form(10),
    cursor: str = Form(""),
    dimensions: int | None = Form(None),
) -> HTMLResponse:
    store, error = _vector_store()
    collections = store.list_collections() if store else []
    page_size = _normalize_k(k)
    offset = 0
    as_of = None
    mode = ""
    query_vector: list[float] | None = None
    cursor_payload: dict[str, Any] | None = None
    query_payload = {
        "collection": collection,
        "text": text,
        "vector": vector,
        "k": page_size,
        "dimensions": dimensions,
    }
    default_collection = _vector_default_collection(collection, collections)
    if error or not store:
        return TEMPLATES.TemplateResponse(
            "vectors.html",
            {
                "request": request,
                "error": error,
                "collections": collections,
                "results": None,
                "query": query_payload,
                "default_collection": default_collection,
                "pagination": None,
            },
            status_code=400,
        )
    try:
        if cursor:
            cursor_payload = _decode_cursor(cursor)
            collection = str(cursor_payload.get("collection") or collection)
            page_size = _normalize_k(cursor_payload.get("page_size"))
            offset = _normalize_offset(cursor_payload.get("offset"))
            as_of = cursor_payload.get("as_of")
            mode = str(cursor_payload.get("mode") or "")
            dimensions = int(cursor_payload.get("dimensions") or 0) or dimensions
            if mode == "text":
                text = str(cursor_payload.get("text") or "")
                query_vector = hash_embedding(text, int(dimensions or 0))
            elif mode == "vector":
                vector_token = str(cursor_payload.get("vector") or "")
                query_vector = _decode_vector(vector_token, int(dimensions or 0))
            else:
                raise ValueError("Invalid cursor mode")
        else:
            if text and vector:
                raise ValueError("Provide either text or vector, not both")
            if not text and not vector:
                raise ValueError("Provide text or vector")
            if vector:
                query_vector = _parse_vector(vector)
                dimensions = len(query_vector)
                mode = "vector"
            else:
                dimensions = _resolve_vector_dimensions(store, collection, dimensions)
                query_vector = hash_embedding(text, int(dimensions))
                mode = "text"
            as_of = now_iso()
            cursor_payload = {
                "v": 1,
                "tenant_id": TENANT_CTX.tenant_id,
                "collection": collection,
                "mode": mode,
                "text": text if mode == "text" else None,
                "vector": _encode_vector(query_vector) if mode == "vector" else None,
                "dimensions": int(dimensions),
                "page_size": page_size,
                "offset": offset,
                "as_of": as_of,
                "embed": "hash_embedding:v1" if mode == "text" else "raw",
            }

        fetch_k = offset + page_size
        if fetch_k > _vector_k_max():
            raise ValueError(
                f"Requested page exceeds max_k={_vector_k_max()}; set STAT_HARNESS_VECTOR_K_MAX"
            )
        results = store.query(
            collection, query_vector, k=fetch_k, as_of=as_of
        )
        page = results[offset : offset + page_size]
        has_next = len(results) > offset + page_size
        has_prev = offset > 0
        next_cursor = None
        prev_cursor = None
        if cursor_payload:
            base_payload = dict(cursor_payload)
            if has_next:
                base_payload["offset"] = offset + page_size
                next_cursor = _encode_cursor(base_payload)
            if has_prev:
                base_payload["offset"] = max(0, offset - page_size)
                prev_cursor = _encode_cursor(base_payload)
        return TEMPLATES.TemplateResponse(
            "vectors.html",
            {
                "request": request,
                "error": None,
                "collections": collections,
                "results": page,
                "query": {
                    "collection": collection,
                    "text": text,
                    "vector": vector,
                    "k": page_size,
                    "dimensions": dimensions,
                },
                "default_collection": default_collection,
                "pagination": {
                    "offset": offset,
                    "page_size": page_size,
                    "has_next": has_next,
                    "has_prev": has_prev,
                    "next_offset": offset + page_size,
                    "prev_offset": max(0, offset - page_size),
                    "cursor": _encode_cursor(cursor_payload) if cursor_payload else None,
                    "next_cursor": next_cursor,
                    "prev_cursor": prev_cursor,
                    "as_of": as_of,
                },
            },
        )
    except Exception as exc:
        return TEMPLATES.TemplateResponse(
            "vectors.html",
            {
                "request": request,
                "error": str(exc),
                "collections": collections,
                "results": None,
                "query": query_payload,
                "default_collection": default_collection,
                "pagination": None,
            },
            status_code=400,
        )


@app.get("/api/vectors/collections")
async def vector_collections_api(request: Request) -> JSONResponse:
    if AUTH_ENABLED and not getattr(request.state, "user", None):
        raise HTTPException(status_code=401, detail="unauthorized")
    store, error = _vector_store()
    if error or not store:
        raise HTTPException(status_code=400, detail=error or "Vector store unavailable")
    return JSONResponse({"collections": store.list_collections()})


@app.post("/api/vectors/query")
async def vector_query_api(request: Request, payload: dict[str, Any]) -> JSONResponse:
    if AUTH_ENABLED and not getattr(request.state, "user", None):
        raise HTTPException(status_code=401, detail="unauthorized")
    store, error = _vector_store()
    if error or not store:
        raise HTTPException(status_code=400, detail=error or "Vector store unavailable")
    cursor = payload.get("cursor")
    collection = str(payload.get("collection") or "")
    text = payload.get("text")
    vector = payload.get("vector")
    page_size = _normalize_k(payload.get("k"))
    offset = _normalize_offset(payload.get("offset"))
    dimensions = payload.get("dimensions")
    as_of = None
    mode = ""
    query_vector: list[float] | None = None
    cursor_payload: dict[str, Any] | None = None
    if cursor:
        cursor_payload = _decode_cursor(str(cursor))
        collection = str(cursor_payload.get("collection") or collection)
        page_size = _normalize_k(cursor_payload.get("page_size"))
        offset = _normalize_offset(cursor_payload.get("offset"))
        as_of = cursor_payload.get("as_of")
        mode = str(cursor_payload.get("mode") or "")
        dimensions = int(cursor_payload.get("dimensions") or 0) or dimensions
        if mode == "text":
            text = str(cursor_payload.get("text") or "")
            query_vector = hash_embedding(text, int(dimensions or 0))
        elif mode == "vector":
            query_vector = _decode_vector(str(cursor_payload.get("vector") or ""), int(dimensions or 0))
        else:
            raise HTTPException(status_code=400, detail="Invalid cursor mode")
    else:
        if not collection:
            raise HTTPException(status_code=400, detail="collection required")
        if text and vector:
            raise HTTPException(status_code=400, detail="Provide either text or vector")
        if not text and not vector:
            raise HTTPException(status_code=400, detail="Provide text or vector")
        if vector is not None:
            if not isinstance(vector, list):
                raise HTTPException(status_code=400, detail="vector must be a list")
            query_vector = [float(val) for val in vector]
            dimensions = len(query_vector)
            mode = "vector"
        else:
            dims = _resolve_vector_dimensions(store, collection, dimensions)
            query_vector = hash_embedding(str(text or ""), dims)
            dimensions = dims
            mode = "text"
        as_of = now_iso()
        cursor_payload = {
            "v": 1,
            "tenant_id": TENANT_CTX.tenant_id,
            "collection": collection,
            "mode": mode,
            "text": text if mode == "text" else None,
            "vector": _encode_vector(query_vector) if mode == "vector" else None,
            "dimensions": int(dimensions),
            "page_size": page_size,
            "offset": offset,
            "as_of": as_of,
            "embed": "hash_embedding:v1" if mode == "text" else "raw",
        }
    try:
        fetch_k = offset + page_size
        if fetch_k > _vector_k_max():
            raise ValueError(
                f"Requested page exceeds max_k={_vector_k_max()}; set STAT_HARNESS_VECTOR_K_MAX"
            )
        results = store.query(collection, query_vector, k=fetch_k, as_of=as_of)
        page = results[offset : offset + page_size]
        next_offset = offset + page_size if len(results) > offset + page_size else None
        next_cursor = None
        prev_cursor = None
        if cursor_payload:
            base_payload = dict(cursor_payload)
            if next_offset is not None:
                base_payload["offset"] = next_offset
                next_cursor = _encode_cursor(base_payload)
            if offset > 0:
                base_payload["offset"] = max(0, offset - page_size)
                prev_cursor = _encode_cursor(base_payload)
        return JSONResponse(
            {
                "results": page,
                "offset": offset,
                "page_size": page_size,
                "next_offset": next_offset,
                "cursor": _encode_cursor(cursor_payload) if cursor_payload else None,
                "next_cursor": next_cursor,
                "prev_cursor": prev_cursor,
                "as_of": as_of,
            }
        )
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc))


def _render_admin(
    request: Request,
    message: str | None = None,
    error: str | None = None,
    new_api_key: str | None = None,
) -> HTMLResponse:
    tenant_id = _current_tenant_id()
    tenants = pipeline.storage.list_tenants()
    users = pipeline.storage.list_users_for_tenant(tenant_id)
    api_keys = pipeline.storage.list_api_keys(tenant_id)
    return TEMPLATES.TemplateResponse(
        "admin.html",
        {
            "request": request,
            "tenant_id": tenant_id,
            "tenants": tenants,
            "users": users,
            "api_keys": api_keys,
            "message": message,
            "error": error,
            "new_api_key": new_api_key,
        },
    )


@app.get("/admin", response_class=HTMLResponse)
async def admin_view(request: Request) -> HTMLResponse:
    _require_admin(request)
    return _render_admin(request)


@app.post("/admin/users")
async def admin_create_user(
    request: Request,
    email: str = Form(...),
    password: str = Form(...),
    name: str = Form(""),
    role: str = Form("member"),
    is_admin: bool = Form(False),
) -> HTMLResponse:
    _require_admin(request)
    normalized = normalize_email(email)
    existing = pipeline.storage.fetch_user_by_email(normalized)
    if existing:
        return _render_admin(request, error="User already exists")
    user_id = pipeline.storage.create_user(
        normalized, hash_password(password), name or None, bool(is_admin), now_iso()
    )
    pipeline.storage.ensure_membership(user_id, role or "member", now_iso())
    return _render_admin(request, message="User created")


@app.post("/admin/users/{user_id}/disable")
async def admin_disable_user(request: Request, user_id: int) -> HTMLResponse:
    _require_admin(request)
    pipeline.storage.disable_user(int(user_id))
    pipeline.storage.revoke_user_sessions(int(user_id))
    pipeline.storage.revoke_user_api_keys(int(user_id))
    return _render_admin(request, message="User disabled")


@app.post("/admin/tenants")
async def admin_create_tenant(
    request: Request,
    tenant_id: str = Form(...),
    name: str = Form(""),
) -> HTMLResponse:
    _require_admin(request)
    tenant_id = tenant_id.strip()
    if not tenant_id:
        return _render_admin(request, error="tenant_id required")
    pipeline.storage.create_tenant(tenant_id, name or None, now_iso())
    current_user = request.state.user
    pipeline.storage.ensure_membership(int(current_user.user_id), "admin", now_iso(), tenant_id)
    return _render_admin(request, message="Tenant created")


@app.post("/admin/api-keys")
async def admin_create_api_key(
    request: Request,
    user_id: int = Form(...),
    name: str = Form(""),
) -> HTMLResponse:
    _require_admin(request)
    token = generate_api_key()
    pipeline.storage.create_api_key(
        int(user_id), hash_token(token), name or None, now_iso()
    )
    return _render_admin(request, message="API key created", new_api_key=token)


@app.post("/admin/api-keys/{key_id}/revoke")
async def admin_revoke_api_key(request: Request, key_id: int) -> HTMLResponse:
    _require_admin(request)
    pipeline.storage.revoke_api_key(int(key_id))
    return _render_admin(request, message="API key revoked")


def _scope_key(scope_type: str, scope_value: str) -> str | None:
    if scope_type == "sha256":
        if pipeline.storage.tenant_id != DEFAULT_TENANT_ID:
            return _safe_sha256(f"{pipeline.storage.tenant_id}:{scope_value}")
        return _safe_sha256(scope_value)
    if not scope_value:
        return None
    if pipeline.storage.tenant_id != DEFAULT_TENANT_ID:
        return scope_key(f"{pipeline.storage.tenant_id}:{scope_type}", scope_value)
    return scope_key(scope_type, scope_value)


def _load_known_issues(scope_type: str, scope_value: str) -> dict[str, object] | None:
    key = _scope_key(scope_type, scope_value)
    if not key:
        return None
    try:
        stored = pipeline.storage.fetch_known_issues(scope_value, scope_type)
    except sqlite3.OperationalError:
        stored = None
    if stored:
        return {
            "strict": bool(stored.get("strict", False)),
            "notes": stored.get("notes") or "",
            "expected_findings": stored.get("expected_findings") or [],
            "natural_language": stored.get("natural_language") or [],
        }
    json_path, _ = _known_issues_paths(key)
    if not json_path.exists():
        return None
    try:
        legacy = json.loads(json_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return None
    if isinstance(legacy, dict):
        normalized = _normalize_known_issues(legacy)
        set_id = pipeline.storage.upsert_known_issue_set(
            scope_value=scope_value,
            scope_type=scope_type,
            upload_id=None,
            strict=bool(normalized.get("strict", False)),
            notes=str(normalized.get("notes") or ""),
            natural_language=normalized.get("natural_language") or [],
        )
        pipeline.storage.replace_known_issues(
            set_id, normalized.get("expected_findings") or []
        )
        return normalized
    return None


def _normalize_known_issues(payload: dict[str, object]) -> dict[str, object]:
    strict = bool(payload.get("strict", False))
    notes = str(payload.get("notes") or "").strip()
    expected = payload.get("expected_findings") or []
    natural = payload.get("natural_language") or payload.get("nl_issues") or []
    cleaned: list[dict[str, object]] = []
    cleaned_nl: list[dict[str, object]] = []

    def clean_kv(data: object) -> dict[str, object]:
        if not isinstance(data, dict):
            return {}
        out: dict[str, object] = {}
        for key, value in data.items():
            if key is None:
                continue
            k = str(key).strip()
            if not k:
                continue
            out[k] = value
        return out

    if isinstance(expected, list):
        for entry in expected:
            if not isinstance(entry, dict):
                continue
            kind = str(entry.get("kind") or "").strip()
            if not kind:
                continue
            item: dict[str, object] = {"kind": kind}
            plugin_id = str(entry.get("plugin_id") or "").strip()
            if plugin_id:
                item["plugin_id"] = plugin_id
            where = clean_kv(entry.get("where"))
            if where:
                item["where"] = where
            contains = clean_kv(entry.get("contains"))
            if contains:
                item["contains"] = contains
            title = str(entry.get("title") or "").strip()
            if title:
                item["title"] = title
            description = str(entry.get("description") or "").strip()
            if description:
                item["description"] = description
            if "min_count" in entry and entry.get("min_count") not in (None, ""):
                try:
                    item["min_count"] = int(entry.get("min_count"))
                except (TypeError, ValueError):
                    pass
            if "max_count" in entry and entry.get("max_count") not in (None, ""):
                try:
                    item["max_count"] = int(entry.get("max_count"))
                except (TypeError, ValueError):
                    pass
            cleaned.append(item)

    if isinstance(natural, list):
        for entry in natural:
            if isinstance(entry, str):
                text = entry.strip()
                if text:
                    cleaned_nl.append({"text": text})
                continue
            if not isinstance(entry, dict):
                continue
            text = str(entry.get("text") or "").strip()
            if not text:
                continue
            item = {"text": text}
            title = str(entry.get("title") or "").strip()
            if title:
                item["title"] = title
            process_hint = str(entry.get("process_hint") or "").strip()
            if process_hint:
                item["process_hint"] = process_hint
            cleaned_nl.append(item)

    return {
        "strict": strict,
        "notes": notes,
        "expected_findings": cleaned,
        "natural_language": cleaned_nl,
    }


def _save_known_issues(
    scope_type: str,
    scope_value: str,
    upload_id: str | None,
    payload: dict[str, object],
) -> dict[str, object]:
    key = _scope_key(scope_type, scope_value)
    if not key:
        raise ValueError("Invalid known-issues scope")
    normalized = _normalize_known_issues(payload)
    compiled, warnings = compile_known_issues(
        normalized.get("natural_language") or []
    )
    combined = list(normalized.get("expected_findings") or [])
    combined.extend(compiled)

    def _dedupe(items: list[dict[str, object]]) -> list[dict[str, object]]:
        seen = set()
        out = []
        for item in items:
            key = (
                item.get("plugin_id"),
                item.get("kind"),
                json_dumps(item.get("where")),
                json_dumps(item.get("contains")),
                item.get("title"),
                item.get("description"),
            )
            if key in seen:
                continue
            seen.add(key)
            out.append(item)
        return out

    combined = _dedupe(combined)
    set_id = pipeline.storage.upsert_known_issue_set(
        scope_value=scope_value,
        scope_type=scope_type,
        upload_id=upload_id,
        strict=bool(normalized.get("strict", False)),
        notes=str(normalized.get("notes") or ""),
        natural_language=normalized.get("natural_language") or [],
    )
    pipeline.storage.replace_known_issues(
        set_id, combined
    )
    json_path, yaml_path = _known_issues_paths(key)
    payload_out = dict(normalized)
    payload_out["expected_findings"] = combined
    payload_out["compiled_count"] = len(compiled)
    payload_out["compile_warnings"] = warnings
    json_path.write_text(json_dumps(payload_out), encoding="utf-8")
    yaml_payload = {
        "strict": payload_out.get("strict", False),
        "notes": payload_out.get("notes", ""),
        "expected_findings": payload_out.get("expected_findings") or [],
    }
    yaml_path.write_text(yaml.safe_dump(yaml_payload, sort_keys=False), encoding="utf-8")
    return payload_out


def _known_issues_yaml(scope_type: str, scope_value: str) -> str | None:
    key = _scope_key(scope_type, scope_value)
    if not key:
        return None
    try:
        stored = pipeline.storage.fetch_known_issues(scope_value, scope_type)
    except sqlite3.OperationalError:
        stored = None
    if stored:
        payload = {
            "strict": bool(stored.get("strict", False)),
            "notes": stored.get("notes") or "",
            "expected_findings": stored.get("expected_findings") or [],
        }
        return yaml.safe_dump(payload, sort_keys=False)
    _, yaml_path = _known_issues_paths(key)
    if yaml_path.exists():
        return yaml_path.read_text(encoding="utf-8")
    return None


def _safe_sha256(value: str) -> str | None:
    if re.fullmatch(r"[a-f0-9]{64}", value or ""):
        return value
    return None


def _slugify(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")
    return slug or uuid.uuid4().hex[:8]


def _unique_project_id(name: str) -> str:
    base = _slugify(name)
    candidate = base
    for idx in range(1, 100):
        if not pipeline.storage.fetch_project(candidate):
            return candidate
        candidate = f"{base}-{idx}"
    return f"{base}-{uuid.uuid4().hex[:4]}"


def _raw_format_fingerprint_for_file(
    path: Path,
    *,
    encoding: str = "utf-8",
    delimiter: str | None = None,
    sheet_name: str | None = None,
    chunk_size: int = 1000,
) -> str | None:
    try:
        import pandas as pd
    except ImportError:
        return None
    suffix = path.suffix.lower()
    columns: list[str] = []
    dtypes: list[object] = []

    try:
        if suffix == ".xlsx":
            import mimetypes

            mimetypes.knownfiles = []
            from openpyxl import load_workbook

            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb[sheet_name] if sheet_name else wb.active
            rows = ws.iter_rows(values_only=True)
            headers = next(rows, None)
            if headers is None:
                return None
            headers = [
                (str(h) if h is not None else f"column_{idx+1}")
                for idx, h in enumerate(headers)
            ]
            sample = []
            for _ in range(int(chunk_size)):
                row = next(rows, None)
                if row is None:
                    break
                sample.append(list(row))
            sample_df = pd.DataFrame(sample, columns=headers)
            columns = list(sample_df.columns)
            dtypes = list(sample_df.dtypes)
        elif suffix == ".json":
            try:
                chunk = next(
                    pd.read_json(path, lines=True, chunksize=int(chunk_size)), None
                )
                if chunk is None:
                    return None
                columns = list(chunk.columns)
                dtypes = list(chunk.dtypes)
            except ValueError:
                df = pd.read_json(path)
                columns = list(df.columns)
                dtypes = list(df.dtypes)
        else:
            chunk = next(
                pd.read_csv(
                    path,
                    delimiter=delimiter,
                    encoding=encoding,
                    chunksize=int(chunk_size),
                ),
                None,
            )
            if chunk is None:
                return None
            columns = list(chunk.columns)
            dtypes = list(chunk.dtypes)
    except Exception:
        return None

    if not columns:
        return None

    fingerprint_payload = [
        {"name": str(col).lower().strip(), "dtype": str(dtype)}
        for col, dtype in zip(columns, dtypes)
    ]
    fingerprint_payload = sorted(fingerprint_payload, key=lambda item: item["name"])
    digest = hashlib.sha256(json_dumps(fingerprint_payload).encode("utf-8")).hexdigest()
    return digest


def _auto_project_for_upload(
    input_file: Path, settings: dict[str, object]
) -> str | None:
    ingest_settings = settings.get("ingest_tabular")
    ingest_settings = ingest_settings if isinstance(ingest_settings, dict) else {}
    encoding = str(ingest_settings.get("encoding") or "utf-8")
    delimiter = ingest_settings.get("delimiter")
    sheet_name = ingest_settings.get("sheet_name")
    try:
        chunk_size = int(ingest_settings.get("chunk_size") or 1000)
    except (TypeError, ValueError):
        chunk_size = 1000

    fingerprint = _raw_format_fingerprint_for_file(
        input_file,
        encoding=encoding,
        delimiter=str(delimiter) if delimiter else None,
        sheet_name=str(sheet_name) if sheet_name else None,
        chunk_size=chunk_size,
    )
    if not fingerprint:
        return None

    raw_format = pipeline.storage.fetch_raw_format_by_fingerprint(fingerprint)
    if raw_format and raw_format.get("format_id") is not None:
        project = pipeline.storage.find_project_for_raw_format(int(raw_format["format_id"]))
        if project and project.get("project_id"):
            return str(project["project_id"])

    name = f"Format {fingerprint[:8]}"
    project_id = _unique_project_id(name)
    pipeline.storage.ensure_project(project_id, project_id, now_iso())
    pipeline.storage.update_project_name(project_id, name)
    pipeline.storage.update_project_erp_type(project_id, "unknown")
    return project_id


def _bool_from_form(value: str | None, default: bool = False) -> bool:
    if value is None:
        return default
    lowered = str(value).strip().lower()
    if lowered in {"1", "true", "yes", "on"}:
        return True
    if lowered in {"0", "false", "no", "off", ""}:
        return False
    return default


def _parse_int(value: str | None, default: int) -> int:
    if value is None or value == "":
        return default
    return int(value)


def _parse_float(value: str | None, default: float) -> float:
    if value is None or value == "":
        return default
    return float(value)


def _format_hours(value: float | int | None) -> str:
    if value is None:
        return "n/a"
    return f"{float(value):.2f}h"


def _format_ratio(value: float | int | None) -> str:
    if value is None:
        return "n/a"
    return f"{float(value):.2f}x"


def _merge_settings(base: dict[str, object], override: dict[str, object]) -> dict[str, object]:
    merged = dict(base)
    for key, value in override.items():
        if isinstance(value, dict) and isinstance(merged.get(key), dict):
            merged[key] = {**merged[key], **value}
        else:
            merged[key] = value
    return merged


def _run_meta_from_plugins(plugin_ids: list[str]) -> object:
    if not plugin_ids:
        return "auto"
    if len(plugin_ids) == 1 and plugin_ids[0] == "all":
        return "all"
    return plugin_ids


def _expected_plugins_for_run(run_row: dict[str, object]) -> list[str]:
    specs = pipeline.manager.discover()
    spec_map = {spec.plugin_id: spec for spec in specs}
    analysis_plugins = [spec.plugin_id for spec in specs if spec.type == "analysis"]
    profile_plugins = [spec.plugin_id for spec in specs if spec.type == "profile"]
    planner_plugins = [spec.plugin_id for spec in specs if spec.type == "planner"]
    transform_plugins = [spec.plugin_id for spec in specs if spec.type == "transform"]
    report_plugins = [spec.plugin_id for spec in specs if spec.type == "report"]
    ingest_plugins = [spec.plugin_id for spec in specs if spec.type == "ingest"]

    settings = {}
    raw_settings = run_row.get("settings_json")
    if raw_settings:
        try:
            settings = json.loads(raw_settings)
        except json.JSONDecodeError:
            settings = {}
    meta = {}
    if isinstance(settings, dict):
        meta = settings.get("__run_meta", {}) if isinstance(settings.get("__run_meta"), dict) else {}
    meta_plugins = meta.get("plugins")

    selected: list[str] = []
    if isinstance(meta_plugins, list):
        selected = [str(pid) for pid in meta_plugins if pid]
    elif isinstance(meta_plugins, str):
        if meta_plugins == "all":
            selected = list(analysis_plugins)
        elif meta_plugins == "auto":
            selected = list(profile_plugins) + list(planner_plugins) + list(analysis_plugins)
        else:
            selected = [meta_plugins]
    else:
        selected = list(analysis_plugins)

    expected: list[str] = []
    input_filename = str(run_row.get("input_filename") or "")
    if not input_filename.startswith("db://"):
        if ingest_plugins:
            expected.append(ingest_plugins[0])
        else:
            expected.append("ingest_tabular")

    for pid in selected:
        if pid in spec_map and pid not in expected:
            expected.append(pid)
        elif pid in {"all", "auto"}:
            continue
        elif pid not in expected:
            expected.append(pid)

    for pid in transform_plugins:
        if pid in selected and pid not in expected:
            expected.append(pid)

    for pid in report_plugins or ["report_bundle"]:
        if pid not in expected:
            expected.append(pid)

    return expected


def _run_label(project: dict | None, run_index: int) -> str:
    project_name = (project or {}).get("name") or (project or {}).get("project_id") or "Project"
    return f"{project_name} Run {run_index}"


def _annotate_runs(
    runs: list[dict[str, object]], project: dict | None
) -> list[dict[str, object]]:
    if not runs:
        return runs
    sorted_runs = sorted(
        runs,
        key=lambda row: row.get("created_at") or "",
    )
    index_by_id = {
        run.get("run_id"): idx + 1
        for idx, run in enumerate(sorted_runs)
    }
    annotated = []
    for run in runs:
        run_id = run.get("run_id")
        label = _run_label(project, index_by_id.get(run_id, 0) or 0)
        copy = dict(run)
        copy["label"] = label
        copy["evaluation"] = _load_evaluation_summary(run_id)
        annotated.append(copy)
    return annotated


def _load_evaluation_summary(run_id: str | None) -> dict[str, object] | None:
    if not run_id:
        return None
    run_dir = APPDATA_DIR / "runs" / str(run_id)
    eval_path = run_dir / "evaluation.json"
    if not eval_path.exists():
        return None
    try:
        payload = json.loads(eval_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return None
    if not isinstance(payload, dict):
        return None
    return {
        "result": payload.get("result"),
        "ok": payload.get("ok"),
        "evaluated_at": payload.get("evaluated_at"),
        "messages": payload.get("messages") if isinstance(payload.get("messages"), list) else [],
    }


def _build_known_issue_results(
    report: dict, known_payload: dict[str, object] | None
) -> dict[str, object] | None:
    if not known_payload or not isinstance(known_payload, dict):
        return None
    expected = known_payload.get("expected_findings") or []
    if not isinstance(expected, list):
        expected = []
    strict = bool(known_payload.get("strict", True))
    results: list[dict[str, object]] = []
    expected_matchers: list[dict[str, object]] = []
    for entry in expected:
        if not isinstance(entry, dict):
            continue
        kind = entry.get("kind")
        if not kind:
            continue
        plugin_id = entry.get("plugin_id") or None
        where = entry.get("where") or {}
        contains = entry.get("contains") or {}
        try:
            min_count = int(entry.get("min_count", 1) or 1)
        except (TypeError, ValueError):
            min_count = 1
        max_count = entry.get("max_count")
        try:
            max_count_val = int(max_count) if max_count is not None else None
        except (TypeError, ValueError):
            max_count_val = None
        candidates = eval_core._collect_findings_for_plugin(
            report, plugin_id, str(kind)
        )
        matches = [
            item
            for item in candidates
            if eval_core._matches_expected(item, where, contains)
        ]
        fallbacks = sorted(
            {
                str(item.get("baseline_match_fallback"))
                for item in matches
                if item.get("baseline_match_fallback")
            }
        )
        baseline_modes = sorted(
            {
                str(item.get("baseline_match_mode"))
                for item in matches
                if item.get("baseline_match_mode")
            }
        )
        baseline_sources = sorted(
            {
                str(item.get("baseline_host_source"))
                for item in matches
                if item.get("baseline_host_source")
            }
        )
        baseline_hosts = sorted(
            {
                str(item.get("baseline_host_count"))
                for item in matches
                if item.get("baseline_host_count") is not None
            }
        )
        status = "pass"
        if len(matches) < min_count:
            status = "fail"
        if max_count_val is not None and len(matches) > max_count_val:
            status = "fail"
        results.append(
            {
                "title": entry.get("title") or str(kind),
                "description": entry.get("description") or "",
                "plugin_id": plugin_id,
                "kind": kind,
                "min_count": min_count,
                "max_count": max_count_val,
                "matched": len(matches),
                "status": status,
                "baseline_match_fallbacks": fallbacks,
                "baseline_match_modes": baseline_modes,
                "baseline_host_sources": baseline_sources,
                "baseline_host_counts": baseline_hosts,
            }
        )
        expected_matchers.append(
            {
                "plugin_id": plugin_id,
                "kind": str(kind),
                "where": where,
                "contains": contains,
            }
        )

    unexpected: list[dict[str, object]] = []
    if strict and expected_matchers:
        seen: set[tuple[str, str]] = set()
        for pid, plugin in report.get("plugins", {}).items():
            for item in plugin.get("findings", []):
                kind = item.get("kind")
                if not kind:
                    continue
                if not any(matcher["kind"] == kind for matcher in expected_matchers):
                    continue
                matched = False
                for matcher in expected_matchers:
                    if matcher["kind"] != kind:
                        continue
                    if matcher["plugin_id"] and matcher["plugin_id"] != pid:
                        continue
                    if eval_core._matches_expected(
                        item, matcher["where"], matcher["contains"]
                    ):
                        matched = True
                        break
                if not matched:
                    key = (str(pid), str(kind))
                    if key not in seen:
                        seen.add(key)
                        unexpected.append({"plugin_id": pid, "kind": kind})

    ok = all(item.get("status") == "pass" for item in results) and (
        not strict or not unexpected
    )
    return {
        "strict": strict,
        "results": results,
        "unexpected": unexpected,
        "ok": ok,
    }


def _build_insights(report: dict) -> dict[str, list[dict[str, object]]]:
    insights: dict[str, list[dict[str, object]]] = {}
    for plugin_id, plugin in report.get("plugins", {}).items():
        entries: list[dict[str, object]] = []
        for finding in plugin.get("findings", []):
            kind = finding.get("kind") or "finding"
            measurement = finding.get("measurement_type", "measured")
            title = f"{kind}"
            bullets: list[str] = []
            if kind == "close_cycle_contention":
                process = finding.get("process") or "process"
                title = f"Close-cycle contention: {process}"
                bullets = [
                    f"Close vs open median duration: {finding.get('median_duration_close')}s vs {finding.get('median_duration_open')}s ({_format_ratio(finding.get('slowdown_ratio'))})",
                    f"Close-cycle runs: {finding.get('close_count')} | Open-cycle runs: {finding.get('open_count')}",
                    f"Correlation (close volume vs median duration): {finding.get('correlation')}",
                    f"Estimated improvement if removed: {float(finding.get('estimated_improvement_pct') or 0) * 100:.1f}%",
                ]
            elif kind == "eligible_wait_process_stats":
                process = finding.get("process") or "process"
                title = f"Eligible wait (standalone): {process}"
                bullets = [
                    f"Total runs: {finding.get('runs_total')} (close: {finding.get('runs_close')}, open: {finding.get('runs_open')})",
                    f"Eligible-wait hours: {_format_hours(finding.get('eligible_wait_hours_total'))} (close: {_format_hours(finding.get('eligible_wait_hours_close'))})",
                    f">threshold eligible-wait: {_format_hours(finding.get('eligible_wait_gt_hours_total'))} (close: {_format_hours(finding.get('eligible_wait_gt_hours_close'))})",
                    f"Pre-eligible wait: {_format_hours(finding.get('wait_pre_hours_total'))}",
                ]
            elif kind == "eligible_wait_impact":
                process = finding.get("process") or "process"
                title = f"Eligible-wait impact without {process}"
                bullets = [
                    f"Total eligible-wait: {_format_hours(finding.get('eligible_wait_hours_total'))} → {_format_hours(finding.get('eligible_wait_hours_without_target'))}",
                    f"Close-window eligible-wait: {_format_hours(finding.get('eligible_wait_hours_close_total'))} → {_format_hours(finding.get('eligible_wait_hours_close_without_target'))}",
                    f"Open-window eligible-wait: {_format_hours(finding.get('eligible_wait_hours_open_total'))} → {_format_hours(finding.get('eligible_wait_hours_open_without_target'))}",
                    f">threshold eligible-wait: {_format_hours(finding.get('eligible_wait_gt_hours_total'))} → {_format_hours(finding.get('eligible_wait_gt_hours_without_target'))}",
                    f"Close-window >threshold: {_format_hours(finding.get('eligible_wait_gt_hours_close_total'))} → {_format_hours(finding.get('eligible_wait_gt_hours_close_without_target'))}",
                    f"Open-window >threshold: {_format_hours(finding.get('eligible_wait_gt_hours_open_total'))} → {_format_hours(finding.get('eligible_wait_gt_hours_open_without_target'))}",
                ]
            elif kind == "capacity_scale_model":
                process = finding.get("process") or "process"
                title = f"Modeled capacity scaling: {process}"
                bullets = [
                    f">threshold eligible-wait (post-removal): {_format_hours(finding.get('eligible_wait_gt_hours_without_target'))}",
                    f"Scale factor: {finding.get('scale_factor')} → modeled {_format_hours(finding.get('eligible_wait_gt_hours_modeled'))}",
                    f"Close-window modeled: {_format_hours(finding.get('eligible_wait_gt_hours_close_without_target'))} → {_format_hours(finding.get('eligible_wait_gt_hours_close_modeled'))}",
                    f"Open-window modeled: {_format_hours(finding.get('eligible_wait_gt_hours_open_without_target'))} → {_format_hours(finding.get('eligible_wait_gt_hours_open_modeled'))}",
                ]
            elif kind == "sequence_classification":
                process = finding.get("process") or "process"
                title = f"Sequence-linked share: {process}"
                bullets = [
                    f"Sequence runs: {finding.get('sequence_runs')} ({float(finding.get('sequence_ratio') or 0) * 100:.1f}%)",
                    f"Standalone runs: {finding.get('standalone_runs')}",
                ]
            elif kind == "process_variant":
                title = "Process sequence variant"
                bullets = [
                    f"Variant length: {len(finding.get('variant') or [])}",
                    f"Count: {finding.get('count')} (fraction: {float(finding.get('fraction') or 0) * 100:.1f}%)",
                ]
            elif kind == "rare_variant":
                title = "Rare process variant"
                bullets = [
                    f"Count: {finding.get('count')}",
                ]

            entries.append(
                {
                    "title": title,
                    "kind": kind,
                    "measurement_type": measurement,
                    "bullets": bullets,
                    "evidence": finding.get("evidence") or {},
                }
            )
        insights[plugin_id] = entries
    return insights


def _normalize_settings_from_form(
    template_name: str,
    lowercase: str | None,
    strip: str | None,
    collapse_whitespace: str | None,
    numeric_coercion: str | None,
    numeric_threshold: str | None,
    exclude_name_patterns: str,
    chunk_size: str | None,
    sample_rows: str | None,
) -> dict[str, object]:
    patterns = [
        token.strip()
        for token in (exclude_name_patterns or "").split(",")
        if token.strip()
    ]
    return {
        "template_name": template_name.strip(),
        "lowercase": _bool_from_form(lowercase, True),
        "strip": _bool_from_form(strip, True),
        "collapse_whitespace": _bool_from_form(collapse_whitespace, True),
        "numeric_coercion": _bool_from_form(numeric_coercion, True),
        "numeric_threshold": _parse_float(numeric_threshold, 0.98),
        "exclude_name_patterns": patterns,
        "chunk_size": _parse_int(chunk_size, 1000),
        "sample_rows": _parse_int(sample_rows, 500),
    }


def _auto_seed(value: str | None, fallback: int = 0) -> int:
    if not value:
        return fallback
    return int(stable_hash(value))


def _load_evaluation_result(run_dir: Path) -> dict[str, object] | None:
    eval_path = run_dir / "evaluation.json"
    if not eval_path.exists():
        return None
    try:
        data = json.loads(eval_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return None
    if not isinstance(data, dict):
        return None
    return data


@app.get("/", response_class=HTMLResponse)
async def index(request: Request) -> HTMLResponse:
    projects = pipeline.storage.list_projects()
    return TEMPLATES.TemplateResponse(
        "index.html", {"request": request, "projects": projects}
    )


@app.get("/known-issues", response_class=HTMLResponse)
async def known_issues(
    request: Request, upload_id: str = "", erp_type: str = "unknown"
) -> HTMLResponse:
    uploads = pipeline.storage.list_uploads()
    selected_id = upload_id or (uploads[0]["upload_id"] if uploads else "")
    upload_row = pipeline.storage.fetch_upload(selected_id) if selected_id else None
    known_payload = None
    known_yaml = ""
    sha256 = None
    scope_type = "erp_type"
    scope_value = (erp_type or "unknown").strip() or "unknown"
    if upload_row and upload_row.get("sha256"):
        sha256 = upload_row["sha256"]
        known_payload = _load_known_issues(scope_type, scope_value)
        known_yaml = _known_issues_yaml(scope_type, scope_value)
    if known_payload is None:
        known_payload = {
            "strict": True,
            "notes": "",
            "expected_findings": [],
            "natural_language": [],
        }
        known_yaml = ""
    specs = pipeline.manager.discover()
    plugin_ids = [spec.plugin_id for spec in specs if spec.type == "analysis"]
    return TEMPLATES.TemplateResponse(
        "known_issues.html",
        {
            "request": request,
            "uploads": uploads,
            "selected_upload": upload_row,
            "selected_upload_id": selected_id,
            "known_issues": known_payload,
            "known_issues_json": json_dumps(known_payload),
            "plugin_ids": plugin_ids,
            "sha256": sha256 or "",
            "known_issues_yaml": known_yaml,
            "erp_type": scope_value,
        },
    )


@app.get("/wizard", response_class=HTMLResponse)
async def wizard(request: Request) -> HTMLResponse:
    uploads = pipeline.storage.list_uploads()
    specs = pipeline.manager.discover()
    plugin_ids = [spec.plugin_id for spec in specs if spec.type == "analysis"]
    return TEMPLATES.TemplateResponse(
        "wizard.html",
        {
            "request": request,
            "uploads": uploads,
            "plugin_ids": plugin_ids,
            "erp_type": "unknown",
        },
    )


@app.get("/plugins", response_class=HTMLResponse)
async def plugins(request: Request) -> HTMLResponse:
    specs = pipeline.manager.discover()
    return TEMPLATES.TemplateResponse(
        "plugins.html", {"request": request, "plugins": specs}
    )


@app.get("/projects", response_class=HTMLResponse)
async def projects_view(request: Request) -> HTMLResponse:
    projects = pipeline.storage.list_projects()
    return TEMPLATES.TemplateResponse(
        "projects.html", {"request": request, "projects": projects}
    )


@app.post("/projects", response_class=HTMLResponse)
async def create_project(
    request: Request, name: str = Form(...), erp_type: str = Form("unknown")
) -> HTMLResponse:
    clean_name = (name or "").strip()
    if not clean_name:
        raise HTTPException(status_code=400, detail="Project name required")
    project_id = _unique_project_id(clean_name)
    pipeline.storage.ensure_project(project_id, project_id, now_iso())
    pipeline.storage.update_project_name(project_id, clean_name)
    if erp_type:
        pipeline.storage.update_project_erp_type(project_id, erp_type.strip() or "unknown")
    return TEMPLATES.TemplateResponse(
        "project.html",
        {
            "request": request,
            "project_id": project_id,
            "datasets": pipeline.storage.list_dataset_versions_by_project(project_id),
            "project": pipeline.storage.fetch_project(project_id),
            "runs": _annotate_runs(
                pipeline.storage.list_runs_by_project(project_id),
                pipeline.storage.fetch_project(project_id),
            ),
            "known_issues": _load_known_issues(
                "erp_type", (erp_type or "unknown").strip() or "unknown"
            )
            or {
                "strict": True,
                "notes": "",
                "expected_findings": [],
                "natural_language": [],
            },
            "message": "Project created.",
        },
    )


@app.get("/projects/{project_id}", response_class=HTMLResponse)
async def project_detail(request: Request, project_id: str) -> HTMLResponse:
    datasets = pipeline.storage.list_dataset_versions_by_project(project_id)
    project = pipeline.storage.fetch_project(project_id)
    runs = _annotate_runs(pipeline.storage.list_runs_by_project(project_id), project)
    known = {
        "strict": True,
        "notes": "",
        "expected_findings": [],
        "natural_language": [],
    }
    if project and project.get("erp_type"):
        known = (
            _load_known_issues("erp_type", project.get("erp_type") or "unknown")
            or known
        )
    plugin_settings = pipeline.storage.fetch_project_plugin_settings(project_id)
    return TEMPLATES.TemplateResponse(
        "project.html",
        {
            "request": request,
            "project_id": project_id,
            "datasets": datasets,
            "project": project,
            "runs": runs,
            "known_issues": known,
            "plugin_settings": plugin_settings,
        },
    )


@app.post("/projects/{project_id}/erp", response_class=HTMLResponse)
async def update_project_erp(
    request: Request, project_id: str, erp_type: str = Form("")
) -> HTMLResponse:
    value = (erp_type or "unknown").strip() or "unknown"
    pipeline.storage.update_project_erp_type(project_id, value)
    datasets = pipeline.storage.list_dataset_versions_by_project(project_id)
    project = pipeline.storage.fetch_project(project_id)
    runs = _annotate_runs(pipeline.storage.list_runs_by_project(project_id), project)
    known = _load_known_issues("erp_type", value) or {
        "strict": True,
        "notes": "",
        "expected_findings": [],
        "natural_language": [],
    }
    return TEMPLATES.TemplateResponse(
        "project.html",
        {
            "request": request,
            "project_id": project_id,
            "datasets": datasets,
            "project": project,
            "runs": runs,
            "known_issues": known,
            "message": "ERP type updated.",
        },
    )


@app.get("/projects/{project_id}/roles", response_class=HTMLResponse)
async def project_roles(
    request: Request, project_id: str, dataset_version_id: str = ""
) -> HTMLResponse:
    datasets = pipeline.storage.list_dataset_versions_by_project(project_id)
    selected = dataset_version_id or (datasets[0]["dataset_version_id"] if datasets else "")
    columns = (
        pipeline.storage.fetch_dataset_columns(selected) if selected else []
    )
    overrides = pipeline.storage.fetch_project_role_overrides(project_id)
    return TEMPLATES.TemplateResponse(
        "project_roles.html",
        {
            "request": request,
            "project_id": project_id,
            "datasets": datasets,
            "dataset_version_id": selected,
            "columns": columns,
            "role_keys": ROLE_KEYS,
            "overrides": overrides,
        },
    )


@app.get("/projects/{project_id}/settings", response_class=HTMLResponse)
async def project_settings_view(request: Request, project_id: str) -> HTMLResponse:
    project = pipeline.storage.fetch_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    specs = pipeline.manager.discover()
    plugin_specs = [spec for spec in specs if spec.type == "analysis"]
    settings = pipeline.storage.fetch_project_plugin_settings(project_id)
    return TEMPLATES.TemplateResponse(
        "project_settings.html",
        {
            "request": request,
            "project_id": project_id,
            "project": project,
            "plugin_specs": plugin_specs,
            "plugin_settings": settings,
        },
    )


@app.post("/projects/{project_id}/settings", response_class=HTMLResponse)
async def project_settings_save(request: Request, project_id: str) -> HTMLResponse:
    project = pipeline.storage.fetch_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    form = await request.form()
    specs = pipeline.manager.discover()
    plugin_specs = [spec for spec in specs if spec.type == "analysis"]
    updated: dict[str, dict[str, object]] = {}
    for spec in plugin_specs:
        field_name = f"settings__{spec.plugin_id}"
        raw = str(form.get(field_name) or "").strip()
        if not raw:
            continue
        try:
            payload = json.loads(raw)
        except json.JSONDecodeError as exc:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid JSON for {spec.plugin_id}: {exc}",
            ) from exc
        if not isinstance(payload, dict):
            raise HTTPException(
                status_code=400,
                detail=f"Settings for {spec.plugin_id} must be a JSON object.",
            )
        updated[spec.plugin_id] = payload

    pipeline.storage.replace_project_plugin_settings(project_id, updated)
    return TEMPLATES.TemplateResponse(
        "project_settings.html",
        {
            "request": request,
            "project_id": project_id,
            "project": project,
            "plugin_specs": plugin_specs,
            "plugin_settings": updated,
            "message": "Settings saved.",
        },
    )


@app.post("/projects/{project_id}/roles", response_class=HTMLResponse)
async def save_project_roles(request: Request, project_id: str) -> HTMLResponse:
    form = await request.form()
    overrides: dict[str, str] = {}
    for role in ROLE_KEYS:
        value = str(form.get(role) or "").strip()
        if value:
            overrides[role] = value
    pipeline.storage.replace_project_role_overrides(project_id, overrides)
    dataset_version_id = str(form.get("dataset_version_id") or "")
    return HTMLResponse(
        "Role overrides saved. "
        f"<a href='/projects/{project_id}/roles?dataset_version_id={dataset_version_id}'>Back</a>"
    )


@app.get("/projects/{project_id}/known-issues", response_class=HTMLResponse)
async def project_known_issues(request: Request, project_id: str) -> HTMLResponse:
    project = pipeline.storage.fetch_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    erp_type = (project.get("erp_type") or "unknown").strip() or "unknown"
    known_payload = _load_known_issues("erp_type", erp_type)
    if known_payload is None:
        known_payload = {
            "strict": True,
            "notes": "",
            "expected_findings": [],
            "natural_language": [],
        }
    specs = pipeline.manager.discover()
    plugin_ids = [spec.plugin_id for spec in specs if spec.type == "analysis"]
    return TEMPLATES.TemplateResponse(
        "project_known_issues.html",
        {
            "request": request,
            "project_id": project_id,
            "project": project,
            "known_issues": known_payload,
            "plugin_ids": plugin_ids,
        },
    )


@app.get("/templates", response_class=HTMLResponse)
async def templates_view(request: Request) -> HTMLResponse:
    templates = pipeline.storage.list_templates()
    return TEMPLATES.TemplateResponse(
        "templates.html", {"request": request, "templates": templates}
    )


@app.get("/templates/{template_id}", response_class=HTMLResponse)
async def template_detail(request: Request, template_id: int) -> HTMLResponse:
    template = pipeline.storage.fetch_template(template_id)
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    fields = pipeline.storage.fetch_template_fields(template_id)
    return TEMPLATES.TemplateResponse(
        "template.html",
        {"request": request, "template": template, "fields": fields},
    )


@app.get("/raw-formats", response_class=HTMLResponse)
async def raw_formats_view(request: Request) -> HTMLResponse:
    formats = pipeline.storage.list_raw_formats()
    return TEMPLATES.TemplateResponse(
        "raw_formats.html", {"request": request, "formats": formats}
    )


@app.get("/raw-formats/{format_id}", response_class=HTMLResponse)
async def raw_format_detail(request: Request, format_id: int) -> HTMLResponse:
    notes = pipeline.storage.list_raw_format_notes(format_id)
    mappings = pipeline.storage.list_raw_format_mappings(format_id)
    templates = pipeline.storage.list_templates()
    return TEMPLATES.TemplateResponse(
        "raw_format.html",
        {
            "request": request,
            "format_id": format_id,
            "notes": notes,
            "mappings": mappings,
            "templates": templates,
        },
    )


@app.post("/api/upload")
async def upload(file: UploadFile = File(...)) -> JSONResponse:
    upload_id = uuid.uuid4().hex
    upload_dir = upload_quarantine_dir(APPDATA_DIR, upload_id)
    upload_dir.mkdir(parents=True, exist_ok=True)
    filename = Path(file.filename).name
    if not filename:
        raise HTTPException(status_code=400, detail="Invalid filename")
    allowed = {".csv", ".tsv", ".txt", ".json", ".xlsx"}
    if Path(filename).suffix.lower() not in allowed:
        raise HTTPException(status_code=400, detail="Unsupported file type")
    target = upload_dir / filename
    hasher = hashlib.sha256()
    total = 0
    max_bytes = max_upload_bytes()
    chunk_size = 8 * 1024 * 1024
    with target.open("wb") as handle:
        while True:
            chunk = await file.read(chunk_size)
            if not chunk:
                break
            total += len(chunk)
            if max_bytes is not None and total > max_bytes:
                handle.flush()
                try:
                    target.unlink()
                except FileNotFoundError:
                    pass
                shutil.rmtree(upload_dir, ignore_errors=True)
                raise HTTPException(status_code=413, detail="File too large")
            hasher.update(chunk)
            handle.write(chunk)
    sha256 = hasher.hexdigest()
    deduplicated = upload_blob_path(APPDATA_DIR, sha256).exists()
    try:
        promote_quarantine_file(
            APPDATA_DIR, upload_id, filename, sha256, verify_on_write=True
        )
    except Exception as exc:
        shutil.rmtree(upload_dir, ignore_errors=True)
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    verified_at = now_iso()
    pipeline.storage.create_upload(
        upload_id, filename, total, sha256, verified_at, verified_at=verified_at
    )
    return JSONResponse(
        {
            "upload_id": upload_id,
            "filename": filename,
            "sha256": sha256,
            "deduplicated": bool(deduplicated),
        }
    )


@app.post("/api/upload/raw")
async def upload_raw(request: Request, filename: str = Query(...)) -> JSONResponse:
    upload_id = uuid.uuid4().hex
    upload_dir = upload_quarantine_dir(APPDATA_DIR, upload_id)
    upload_dir.mkdir(parents=True, exist_ok=True)
    filename = Path(filename).name
    if not filename:
        raise HTTPException(status_code=400, detail="Invalid filename")
    allowed = {".csv", ".tsv", ".txt", ".json", ".xlsx"}
    if Path(filename).suffix.lower() not in allowed:
        raise HTTPException(status_code=400, detail="Unsupported file type")
    target = upload_dir / filename
    hasher = hashlib.sha256()
    total = 0
    max_bytes = max_upload_bytes()
    chunk_size = 8 * 1024 * 1024
    with target.open("wb") as handle:
        async for chunk in request.stream():
            if not chunk:
                continue
            total += len(chunk)
            if max_bytes is not None and total > max_bytes:
                handle.flush()
                try:
                    target.unlink()
                except FileNotFoundError:
                    pass
                shutil.rmtree(upload_dir, ignore_errors=True)
                raise HTTPException(status_code=413, detail="File too large")
            hasher.update(chunk)
            handle.write(chunk)
            if total % (chunk_size * 16) == 0:
                handle.flush()
    sha256 = hasher.hexdigest()
    deduplicated = upload_blob_path(APPDATA_DIR, sha256).exists()
    try:
        promote_quarantine_file(
            APPDATA_DIR, upload_id, filename, sha256, verify_on_write=True
        )
    except Exception as exc:
        shutil.rmtree(upload_dir, ignore_errors=True)
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    verified_at = now_iso()
    pipeline.storage.create_upload(
        upload_id, filename, total, sha256, verified_at, verified_at=verified_at
    )
    return JSONResponse(
        {
            "upload_id": upload_id,
            "filename": filename,
            "sha256": sha256,
            "deduplicated": bool(deduplicated),
        }
    )


@app.post("/api/runs")
async def create_run(
    background: BackgroundTasks,
    upload_id: str = Form(...),
    project_id: str = Form(""),
    plugins: str = Form("all"),
    settings_json: str = Form(""),
    run_seed: int | None = Form(None),
    normalize_template_name: str = Form(""),
    normalize_lowercase: str | None = Form(None),
    normalize_strip: str | None = Form(None),
    normalize_collapse_whitespace: str | None = Form(None),
    normalize_numeric_coercion: str | None = Form(None),
    normalize_numeric_threshold: str = Form("0.98"),
    normalize_exclude_name_patterns: str = Form("id,uuid,guid,key"),
    normalize_chunk_size: str = Form("1000"),
    normalize_sample_rows: str = Form("500"),
) -> JSONResponse:
    upload_row = pipeline.storage.fetch_upload(upload_id)
    if not upload_row:
        raise HTTPException(status_code=404, detail="Upload not found")
    sha256 = str(upload_row.get("sha256") or "")
    input_file = upload_blob_path(APPDATA_DIR, sha256) if sha256 else None
    if not input_file or not input_file.exists():
        # Legacy path fallback.
        upload_dir = APPDATA_DIR / "uploads" / upload_id
        if not upload_dir.exists():
            raise HTTPException(status_code=404, detail="Upload file not found")
        files = list(upload_dir.iterdir())
        if not files:
            raise HTTPException(status_code=400, detail="No uploaded file")
        input_file = files[0]
    plugin_ids = [p for p in plugins.split(",") if p]
    settings: dict[str, object] = {}
    if settings_json:
        try:
            settings = json.loads(settings_json)
        except json.JSONDecodeError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
    if not isinstance(settings, dict):
        raise HTTPException(status_code=400, detail="settings_json must be an object")

    project_id = project_id.strip()
    if not project_id:
        project_id = _auto_project_for_upload(input_file, settings) or ""

    project_settings = {}
    if project_id:
        project_settings = pipeline.storage.fetch_project_plugin_settings(project_id)

    normalize_settings = _normalize_settings_from_form(
        normalize_template_name,
        normalize_lowercase,
        normalize_strip,
        normalize_collapse_whitespace,
        normalize_numeric_coercion,
        normalize_numeric_threshold,
        normalize_exclude_name_patterns,
        normalize_chunk_size,
        normalize_sample_rows,
    )
    plugin_settings = dict(settings.get("transform_normalize_mixed", {}))
    plugin_settings.update(normalize_settings)
    settings["transform_normalize_mixed"] = plugin_settings
    if project_settings:
        settings = _merge_settings(project_settings, settings)
    settings["__run_meta"] = {"plugins": _run_meta_from_plugins(plugin_ids)}

    if run_seed in (None, 0):
        run_seed = _auto_seed(upload_row.get("sha256") or upload_id, 0)

    run_id = uuid.uuid4().hex

    def run_pipeline() -> None:
        pipeline.run(
            input_file,
            plugin_ids,
            settings,
            int(run_seed or 0),
            run_id=run_id,
            upload_id=upload_id,
            project_id=project_id or None,
        )
        run_dir = APPDATA_DIR / "runs" / run_id
        report = build_report(
            pipeline.storage, run_id, run_dir, Path("docs/report.schema.json")
        )
        write_report(report, run_dir)
        project_row = (
            pipeline.storage.fetch_project(project_id)
            if project_id
            else None
        )
        ground_truth, _source = _resolve_ground_truth(
            "known", report, upload_row, project_row, None
        )
        if ground_truth:
            gt_path = run_dir / "ground_truth.yaml"
            gt_path.write_text(ground_truth, encoding="utf-8")
            ok, messages = evaluate_report(run_dir / "report.json", gt_path)
            _write_evaluation(run_dir, ok, messages)

    background.add_task(run_pipeline)
    return JSONResponse({"status": "queued", "run_id": run_id})


@app.post("/runs/auto-evaluate", response_class=HTMLResponse)
async def run_auto_evaluate(
    request: Request,
    background: BackgroundTasks,
    upload_id: str = Form(...),
    project_id: str = Form(""),
    run_seed: int | None = Form(None),
    normalize_template_name: str = Form(""),
    normalize_lowercase: str | None = Form(None),
    normalize_strip: str | None = Form(None),
    normalize_collapse_whitespace: str | None = Form(None),
    normalize_numeric_coercion: str | None = Form(None),
    normalize_numeric_threshold: str = Form("0.98"),
    normalize_exclude_name_patterns: str = Form("id,uuid,guid,key"),
    normalize_chunk_size: str = Form("1000"),
    normalize_sample_rows: str = Form("500"),
) -> HTMLResponse:
    upload_row = pipeline.storage.fetch_upload(upload_id)
    if not upload_row:
        raise HTTPException(status_code=404, detail="Upload not found")
    sha256 = str(upload_row.get("sha256") or "")
    input_file = upload_blob_path(APPDATA_DIR, sha256) if sha256 else None
    if not input_file or not input_file.exists():
        # Legacy path fallback.
        upload_dir = APPDATA_DIR / "uploads" / upload_id
        if not upload_dir.exists():
            raise HTTPException(status_code=404, detail="Upload file not found")
        files = list(upload_dir.iterdir())
        if not files:
            raise HTTPException(status_code=400, detail="No uploaded file")
        input_file = files[0]

    if run_seed in (None, 0):
        run_seed = _auto_seed(upload_row.get("sha256") or upload_id, 0)

    settings: dict[str, object] = {}
    project_id = project_id.strip()
    if not project_id:
        project_id = _auto_project_for_upload(input_file, settings) or ""
    project_settings = {}
    if project_id:
        project_settings = pipeline.storage.fetch_project_plugin_settings(project_id)
    normalize_settings = _normalize_settings_from_form(
        normalize_template_name,
        normalize_lowercase,
        normalize_strip,
        normalize_collapse_whitespace,
        normalize_numeric_coercion,
        normalize_numeric_threshold,
        normalize_exclude_name_patterns,
        normalize_chunk_size,
        normalize_sample_rows,
    )
    settings["transform_normalize_mixed"] = normalize_settings
    if project_settings:
        settings = _merge_settings(project_settings, settings)
    settings["__run_meta"] = {"plugins": _run_meta_from_plugins([])}

    run_id = uuid.uuid4().hex

    def run_pipeline() -> None:
        pipeline.run(
            input_file,
            [],
            settings,
            int(run_seed or 0),
            run_id=run_id,
            upload_id=upload_id,
            project_id=project_id or None,
        )
        run_dir = APPDATA_DIR / "runs" / run_id
        report = build_report(
            pipeline.storage, run_id, run_dir, Path("docs/report.schema.json")
        )
        write_report(report, run_dir)

    background.add_task(run_pipeline)
    return TEMPLATES.TemplateResponse(
        "auto_evaluate.html",
        {
            "request": request,
            "run_id": run_id,
        },
    )


@app.post("/api/runs/auto-evaluate")
async def api_auto_evaluate(
    background: BackgroundTasks,
    upload_id: str = Form(...),
    project_id: str = Form(""),
    run_seed: int | None = Form(None),
    normalize_template_name: str = Form(""),
    normalize_lowercase: str | None = Form(None),
    normalize_strip: str | None = Form(None),
    normalize_collapse_whitespace: str | None = Form(None),
    normalize_numeric_coercion: str | None = Form(None),
    normalize_numeric_threshold: str = Form("0.98"),
    normalize_exclude_name_patterns: str = Form("id,uuid,guid,key"),
    normalize_chunk_size: str = Form("1000"),
    normalize_sample_rows: str = Form("500"),
) -> JSONResponse:
    upload_row = pipeline.storage.fetch_upload(upload_id)
    if not upload_row:
        raise HTTPException(status_code=404, detail="Upload not found")
    sha256 = str(upload_row.get("sha256") or "")
    input_file = upload_blob_path(APPDATA_DIR, sha256) if sha256 else None
    if not input_file or not input_file.exists():
        # Legacy path fallback.
        upload_dir = APPDATA_DIR / "uploads" / upload_id
        if not upload_dir.exists():
            raise HTTPException(status_code=404, detail="Upload file not found")
        files = list(upload_dir.iterdir())
        if not files:
            raise HTTPException(status_code=400, detail="No uploaded file")
        input_file = files[0]

    if run_seed in (None, 0):
        run_seed = _auto_seed(upload_row.get("sha256") or upload_id, 0)

    settings: dict[str, object] = {}
    project_id = project_id.strip()
    if not project_id:
        project_id = _auto_project_for_upload(input_file, settings) or ""
    project_settings = {}
    if project_id:
        project_settings = pipeline.storage.fetch_project_plugin_settings(project_id)
    normalize_settings = _normalize_settings_from_form(
        normalize_template_name,
        normalize_lowercase,
        normalize_strip,
        normalize_collapse_whitespace,
        normalize_numeric_coercion,
        normalize_numeric_threshold,
        normalize_exclude_name_patterns,
        normalize_chunk_size,
        normalize_sample_rows,
    )
    settings["transform_normalize_mixed"] = normalize_settings
    if project_settings:
        settings = _merge_settings(project_settings, settings)
    settings["__run_meta"] = {"plugins": _run_meta_from_plugins([])}

    run_id = uuid.uuid4().hex

    def run_pipeline() -> None:
        pipeline.run(
            input_file,
            [],
            settings,
            int(run_seed or 0),
            run_id=run_id,
            upload_id=upload_id,
            project_id=project_id or None,
        )
        run_dir = APPDATA_DIR / "runs" / run_id
        report = build_report(
            pipeline.storage, run_id, run_dir, Path("docs/report.schema.json")
        )
        write_report(report, run_dir)

    background.add_task(run_pipeline)
    return JSONResponse({"status": "queued", "run_id": run_id})


@app.get("/api/projects")
async def list_projects() -> JSONResponse:
    return JSONResponse({"projects": pipeline.storage.list_projects()})


@app.get("/api/uploads")
async def list_uploads(limit: int = 50) -> JSONResponse:
    return JSONResponse({"uploads": pipeline.storage.list_uploads(limit=limit)})


@app.get("/api/known-issues")
async def get_known_issues(
    upload_id: str | None = None,
    sha256: str | None = None,
    erp_type: str | None = None,
) -> JSONResponse:
    scope_type = "sha256"
    scope_value = ""
    if erp_type:
        scope_type = "erp_type"
        scope_value = erp_type.strip() or "unknown"
    elif sha256:
        scope_value = sha256
    elif upload_id:
        upload_row = pipeline.storage.fetch_upload(str(upload_id))
        if not upload_row:
            raise HTTPException(status_code=404, detail="Upload not found")
        scope_value = upload_row.get("sha256") or ""

    key = _scope_key(scope_type, scope_value)
    if not key:
        raise HTTPException(status_code=400, detail="Invalid known-issues scope")
    known_payload = _load_known_issues(scope_type, scope_value)
    if known_payload is None:
        known_payload = {
            "strict": True,
            "notes": "",
            "expected_findings": [],
            "natural_language": [],
        }
    return JSONResponse(
        {
            "sha256": key,
            "scope_type": scope_type,
            "scope_value": scope_value,
            "known_issues": known_payload,
            "ground_truth_yaml": _known_issues_yaml(scope_type, scope_value) or "",
        }
    )


@app.post("/api/known-issues")
async def save_known_issues(payload: dict) -> JSONResponse:
    upload_id = payload.get("upload_id")
    if not upload_id:
        raise HTTPException(status_code=400, detail="upload_id required")
    upload_row = pipeline.storage.fetch_upload(str(upload_id))
    if not upload_row:
        raise HTTPException(status_code=404, detail="Upload not found")
    erp_type = str(payload.get("erp_type") or "").strip()
    scope_type = "erp_type" if erp_type else "sha256"
    scope_value = erp_type or (upload_row.get("sha256") or "")
    known = payload.get("known_issues") or payload.get("payload") or {}
    if not isinstance(known, dict):
        raise HTTPException(status_code=400, detail="known_issues must be an object")
    try:
        normalized = _save_known_issues(scope_type, scope_value, str(upload_id), known)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    key = _scope_key(scope_type, scope_value) or ""
    return JSONResponse(
        {
            "status": "ok",
            "upload_id": upload_id,
            "sha256": key,
            "scope_type": scope_type,
            "scope_value": scope_value,
            "known_issues": normalized,
            "ground_truth_yaml": _known_issues_yaml(scope_type, scope_value) or "",
        }
    )


@app.get("/api/projects/{project_id}")
async def get_project(project_id: str) -> JSONResponse:
    datasets = pipeline.storage.list_dataset_versions_by_project(project_id)
    project = pipeline.storage.fetch_project(project_id)
    return JSONResponse({"project": project, "datasets": datasets})


@app.get("/api/projects/{project_id}/known-issues")
async def get_project_known_issues(project_id: str) -> JSONResponse:
    project = pipeline.storage.fetch_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    erp_type = (project.get("erp_type") or "unknown").strip() or "unknown"
    known_payload = _load_known_issues("erp_type", erp_type)
    if known_payload is None:
        known_payload = {
            "strict": True,
            "notes": "",
            "expected_findings": [],
            "natural_language": [],
        }
    return JSONResponse(
        {
            "project_id": project_id,
            "erp_type": erp_type,
            "known_issues": known_payload,
        }
    )


@app.post("/api/projects/{project_id}/known-issues")
async def save_project_known_issues(project_id: str, payload: dict) -> JSONResponse:
    project = pipeline.storage.fetch_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    erp_type = (project.get("erp_type") or "unknown").strip() or "unknown"
    known = payload.get("known_issues") or payload.get("payload") or {}
    if not isinstance(known, dict):
        raise HTTPException(status_code=400, detail="known_issues must be an object")
    normalized = _save_known_issues("erp_type", erp_type, None, known)
    return JSONResponse(
        {
            "status": "ok",
            "project_id": project_id,
            "erp_type": erp_type,
            "known_issues": normalized,
        }
    )


@app.get("/api/templates")
async def list_templates() -> JSONResponse:
    return JSONResponse({"templates": pipeline.storage.list_templates()})


@app.post("/api/templates")
async def create_template(payload: dict) -> JSONResponse:
    name = payload.get("name")
    if not name:
        raise HTTPException(status_code=400, detail="Template name required")
    fields = payload.get("fields") or []
    if not isinstance(fields, list) or not fields:
        raise HTTPException(status_code=400, detail="Template fields required")
    template_id = pipeline.storage.create_template(
        name=name,
        fields=fields,
        description=payload.get("description"),
        version=payload.get("version"),
        created_at=now_iso(),
    )
    return JSONResponse({"template_id": template_id})


@app.post("/templates/{template_id}/map", response_class=HTMLResponse)
async def map_template(
    request: Request,
    template_id: int,
    dataset_version_id: str = Form(...),
    mapping_json: str = Form(...),
) -> HTMLResponse:
    try:
        mapping = json.loads(mapping_json)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    settings = {"transform_template": {"template_id": int(template_id), "mapping": mapping}}
    run_id = pipeline.run(
        None,
        ["transform_template"],
        settings,
        0,
        dataset_version_id=dataset_version_id,
    )
    return HTMLResponse(
        f"Template mapping queued: {run_id}. <a href='/runs/{run_id}'>View</a>"
    )


@app.get("/api/templates/{template_id}")
async def get_template(template_id: int) -> JSONResponse:
    template = pipeline.storage.fetch_template(template_id)
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    fields = pipeline.storage.fetch_template_fields(template_id)
    return JSONResponse({"template": template, "fields": fields})


@app.post("/templates/{template_id}/run", response_class=HTMLResponse)
async def run_template_combined(
    request: Request,
    template_id: int,
    plugins: str = Form(""),
    settings_json: str = Form(""),
    run_seed: int | None = Form(None),
    project_ids: str = Form(""),
    dataset_ids: str = Form(""),
    dataset_version_ids: str = Form(""),
    raw_format_ids: str = Form(""),
    created_after: str = Form(""),
    created_before: str = Form(""),
) -> HTMLResponse:
    plugin_ids = [p for p in plugins.split(",") if p]
    settings = {}
    if settings_json:
        try:
            settings = json.loads(settings_json)
        except json.JSONDecodeError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    filters = _parse_filter_params(
        project_ids,
        dataset_ids,
        dataset_version_ids,
        raw_format_ids,
        created_after,
        created_before,
    )

    dataset_version_id = pipeline.storage.ensure_template_aggregate_dataset(
        int(template_id), now_iso(), filters=filters or None
    )
    if run_seed in (None, 0):
        run_seed = _auto_seed(dataset_version_id, 0)
    run_id = pipeline.run(
        None,
        plugin_ids,
        settings,
        int(run_seed or 0),
        dataset_version_id=dataset_version_id,
    )
    return HTMLResponse(
        f"Run started: {run_id}. <a href='/runs/{run_id}'>View</a>"
    )


@app.get("/api/trace/row")
async def trace_row_api(
    dataset_version_id: str,
    row_index: int,
    source_dataset_version_id: str | None = None,
    max_rows: int = 50,
) -> JSONResponse:
    payload = pipeline.storage.fetch_row_trace(
        dataset_version_id,
        int(row_index),
        source_dataset_version_id=source_dataset_version_id,
        max_rows=max_rows,
    )
    return JSONResponse(payload)


@app.get("/trace/row", response_class=HTMLResponse)
async def trace_row_view(
    request: Request,
    dataset_version_id: str,
    row_index: int,
    source_dataset_version_id: str | None = None,
    max_rows: int = 50,
) -> HTMLResponse:
    payload = pipeline.storage.fetch_row_trace(
        dataset_version_id,
        int(row_index),
        source_dataset_version_id=source_dataset_version_id,
        max_rows=max_rows,
    )
    return TEMPLATES.TemplateResponse(
        "row_trace.html",
        {
            "request": request,
            "payload": payload,
            "dataset_version_id": dataset_version_id,
            "row_index": row_index,
            "source_dataset_version_id": source_dataset_version_id or "",
        },
    )


@app.get("/templates/{template_id}/results", response_class=HTMLResponse)
async def template_results(request: Request, template_id: int) -> HTMLResponse:
    filters = _parse_filter_params(
        request.query_params.get("project_ids", ""),
        request.query_params.get("dataset_ids", ""),
        request.query_params.get("dataset_version_ids", ""),
        request.query_params.get("raw_format_ids", ""),
        request.query_params.get("created_after", ""),
        request.query_params.get("created_before", ""),
    )
    dataset_version_id = pipeline.storage.ensure_template_aggregate_dataset(
        int(template_id), now_iso(), filters=filters or None
    )
    results = pipeline.storage.fetch_latest_plugin_results_for_dataset(
        dataset_version_id
    )
    return TEMPLATES.TemplateResponse(
        "template_results.html",
        {
            "request": request,
            "template_id": template_id,
            "dataset_version_id": dataset_version_id,
            "results": results,
            "filters": filters,
            "filter_params": {
                "project_ids": request.query_params.get("project_ids", ""),
                "dataset_ids": request.query_params.get("dataset_ids", ""),
                "dataset_version_ids": request.query_params.get("dataset_version_ids", ""),
                "raw_format_ids": request.query_params.get("raw_format_ids", ""),
                "created_after": request.query_params.get("created_after", ""),
                "created_before": request.query_params.get("created_before", ""),
            },
        },
    )


@app.post("/api/datasets/{dataset_version_id}/template")
async def set_dataset_template(dataset_version_id: str, payload: dict) -> JSONResponse:
    template_id = payload.get("template_id")
    mapping = payload.get("mapping") or {}
    if not template_id:
        raise HTTPException(status_code=400, detail="template_id required")
    if not isinstance(mapping, dict):
        raise HTTPException(status_code=400, detail="mapping must be object")
    settings = {"transform_template": {"template_id": int(template_id), "mapping": mapping}}
    run_id = pipeline.run(
        None,
        ["transform_template"],
        settings,
        0,
        dataset_version_id=dataset_version_id,
    )
    return JSONResponse(
        {"status": "queued", "template_id": template_id, "run_id": run_id}
    )


@app.get("/api/raw-formats")
async def list_raw_formats() -> JSONResponse:
    return JSONResponse({"formats": pipeline.storage.list_raw_formats()})


@app.post("/api/raw-formats/{format_id}/notes")
async def add_raw_format_note(format_id: int, payload: dict) -> JSONResponse:
    note = payload.get("note")
    if not note:
        raise HTTPException(status_code=400, detail="note required")
    pipeline.storage.add_raw_format_note(format_id, note, now_iso())
    return JSONResponse({"status": "ok"})


@app.post("/raw-formats/{format_id}/notes", response_class=HTMLResponse)
async def add_raw_format_note_form(
    request: Request, format_id: int, note: str = Form(...)
) -> HTMLResponse:
    pipeline.storage.add_raw_format_note(format_id, note, now_iso())
    return HTMLResponse("Note added. <a href='/raw-formats'>Back</a>")


@app.get("/api/raw-formats/{format_id}/notes")
async def list_raw_format_notes(format_id: int) -> JSONResponse:
    return JSONResponse({"notes": pipeline.storage.list_raw_format_notes(format_id)})


@app.get("/api/raw-formats/{format_id}/mappings")
async def list_raw_format_mappings(format_id: int) -> JSONResponse:
    return JSONResponse(
        {"mappings": pipeline.storage.list_raw_format_mappings(format_id)}
    )


@app.post("/raw-formats/{format_id}/mappings", response_class=HTMLResponse)
async def add_raw_format_mapping_form(
    request: Request,
    format_id: int,
    template_id: int = Form(...),
    mapping_json: str = Form(...),
    notes: str = Form(""),
) -> HTMLResponse:
    try:
        mapping = json.loads(mapping_json)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    mapping_hash = hashlib.sha256(json_dumps(mapping).encode("utf-8")).hexdigest()
    pipeline.storage.add_raw_format_mapping(
        format_id,
        int(template_id),
        json_dumps(mapping),
        mapping_hash,
        notes or None,
        now_iso(),
    )
    return HTMLResponse("Mapping saved. <a href='/raw-formats'>Back</a>")



def _write_evaluation(run_dir: Path, ok: bool, messages: list[str]) -> dict[str, object]:
    payload = {
        "evaluated_at": now_iso(),
        "result": "passed" if ok else "failed",
        "ok": bool(ok),
        "messages": messages,
    }
    eval_path = run_dir / "evaluation.json"
    atomic_write_text(eval_path, json_dumps(payload) + "\n")
    log_path = run_dir / "logs" / "evaluate.log"
    log_path.parent.mkdir(parents=True, exist_ok=True)
    with log_path.open("a", encoding="utf-8") as handle:
        handle.write(f"[{payload['evaluated_at']}] {payload['result']}\n")
        for message in messages:
            handle.write(f"- {message}\n")
    run_log = run_dir / "logs" / "run.log"
    with run_log.open("a", encoding="utf-8") as handle:
        handle.write(f"[{payload['evaluated_at']}] evaluation {payload['result']}\n")
        for message in messages:
            handle.write(f"- {message}\n")
    run_id = run_dir.name
    run_row = pipeline.storage.fetch_run(run_id)
    if run_row:
        if not ok:
            pipeline.storage.update_run_status(
                run_id,
                "failed",
                error={"type": "KnownIssuesFailed", "messages": messages},
            )
        else:
            # Don't override partial/completed statuses written by the pipeline.
            if run_row.get("status") in {"failed", "error", "aborted"}:
                pipeline.storage.update_run_status(run_id, "completed", error=None)
    return payload

def _resolve_ground_truth(
    mode: str,
    report: dict,
    upload_row: dict | None,
    project_row: dict | None,
    provided: str | None,
) -> tuple[str, str]:
    if provided:
        return provided, "manual"
    mode = (mode or "auto").lower()
    if mode in {"known", "auto"}:
        if project_row and project_row.get("erp_type"):
            erp_type = str(project_row.get("erp_type") or "unknown").strip() or "unknown"
            known = _known_issues_yaml("erp_type", erp_type)
            if known:
                return known, "known"
        if upload_row and upload_row.get("sha256"):
            sha256 = _safe_sha256(upload_row.get("sha256") or "")
            if sha256:
                known = _known_issues_yaml("sha256", sha256)
                if known:
                    return known, "known"
    if mode in {"template", "auto"}:
        return _ground_truth_template(report), "template"
    return "", ""


def _plugin_code_hash(spec) -> str | None:
    module_path, _ = spec.entrypoint.split(":", 1)
    if module_path.endswith(".py"):
        module_file = spec.path / module_path
    else:
        module_file = spec.path / f"{module_path}.py"
    return file_sha256(module_file) if module_file.exists() else None


@app.get("/api/datasets/{dataset_version_id}/status")
async def dataset_status(dataset_version_id: str) -> JSONResponse:
    specs = [s for s in pipeline.manager.discover() if s.type == "analysis"]
    dataset = pipeline.storage.get_dataset_version(dataset_version_id)
    if not dataset:
        raise HTTPException(status_code=404, detail="Dataset not found")
    dataset_hash = dataset.get("data_hash")
    results = pipeline.storage.fetch_latest_plugin_results_for_dataset(
        dataset_version_id
    )
    deliveries = pipeline.storage.fetch_deliveries_for_dataset(dataset_version_id)
    results_map = {row["plugin_id"]: row for row in results}
    delivery_map = {}
    for row in deliveries:
        if row["plugin_id"] not in delivery_map:
            delivery_map[row["plugin_id"]] = row

    status_rows = []
    for spec in specs:
        current_hash = _plugin_code_hash(spec)
        result = results_map.get(spec.plugin_id)
        delivery = delivery_map.get(spec.plugin_id)
        status = "pending"
        if result:
            status = "ran"
        if delivery:
            if delivery.get("code_hash") == current_hash and delivery.get(
                "dataset_hash"
            ) == dataset_hash:
                status = "delivered"
            else:
                status = "stale"
        if not result:
            status = "pending"
        if result and not delivery:
            status = "needs_delivery"
        status_rows.append(
            {
                "plugin_id": spec.plugin_id,
                "plugin_version": spec.version,
                "executed_at": result.get("executed_at") if result else None,
                "delivered_at": delivery.get("delivered_at") if delivery else None,
                "status": status,
            }
        )
    return JSONResponse({"dataset_version_id": dataset_version_id, "plugins": status_rows})


@app.post("/api/datasets/{dataset_version_id}/rerun")
async def rerun_dataset(
    background: BackgroundTasks, dataset_version_id: str
) -> JSONResponse:
    ctx = pipeline.storage.get_dataset_version_context(dataset_version_id)
    if not ctx:
        raise HTTPException(status_code=404, detail="Dataset not found")
    project_id = ctx.get("project_id")
    project_settings = (
        pipeline.storage.fetch_project_plugin_settings(project_id)
        if project_id
        else {}
    )
    settings: dict[str, object] = {}
    if project_settings:
        settings = _merge_settings(project_settings, settings)
    settings["__run_meta"] = {"plugins": _run_meta_from_plugins(["all"])}
    run_seed = _auto_seed(dataset_version_id, 0)
    run_id = uuid.uuid4().hex

    def run_pipeline() -> None:
        pipeline.run(
            None,
            ["all"],
            settings,
            int(run_seed or 0),
            run_id=run_id,
            dataset_version_id=dataset_version_id,
            project_id=project_id,
        )
        run_dir = APPDATA_DIR / "runs" / run_id
        report = build_report(
            pipeline.storage, run_id, run_dir, Path("docs/report.schema.json")
        )
        write_report(report, run_dir)
        project_row = (
            pipeline.storage.fetch_project(project_id)
            if project_id
            else None
        )
        ground_truth, _source = _resolve_ground_truth(
            "known", report, None, project_row, None
        )
        if ground_truth:
            gt_path = run_dir / "ground_truth.yaml"
            gt_path.write_text(ground_truth, encoding="utf-8")
            ok, messages = evaluate_report(run_dir / "report.json", gt_path)
            _write_evaluation(run_dir, ok, messages)

    background.add_task(run_pipeline)
    return JSONResponse({"status": "queued", "run_id": run_id})


@app.get("/api/runs/{run_id}/progress")
async def run_progress(run_id: str) -> JSONResponse:
    run_row = pipeline.storage.fetch_run(run_id)
    if not run_row:
        raise HTTPException(status_code=404, detail="Run not found")
    expected = _expected_plugins_for_run(run_row)
    executions = pipeline.storage.fetch_plugin_executions(run_id)
    latest: dict[str, dict[str, object]] = {}
    for row in executions:
        latest[str(row.get("plugin_id"))] = row

    plugin_statuses = []
    completed = 0
    for plugin_id in expected:
        entry = latest.get(plugin_id)
        status = "pending"
        started_at = None
        completed_at = None
        if entry:
            started_at = entry.get("started_at")
            completed_at = entry.get("completed_at")
            status = str(entry.get("status") or "running")
            if completed_at is None:
                status = "running"
        else:
            if run_row.get("status") == "completed":
                status = "not_run"
        if status not in {"pending", "running"}:
            completed += 1
        plugin_statuses.append(
            {
                "plugin_id": plugin_id,
                "status": status,
                "started_at": started_at,
                "completed_at": completed_at,
            }
        )

    total = len(expected) if expected else len(executions)
    percent = int((completed / total) * 100) if total else 0
    errors = []
    for row in pipeline.storage.fetch_plugin_results(run_id):
        if row.get("status") != "error" and not row.get("error_json"):
            continue
        message = row.get("summary") or "Plugin error"
        err_payload = {}
        if row.get("error_json"):
            try:
                err_payload = json.loads(row["error_json"])
            except json.JSONDecodeError:
                err_payload = {}
        if err_payload.get("message"):
            message = err_payload["message"]
        errors.append(
            {
                "plugin_id": row.get("plugin_id"),
                "message": message,
                "type": err_payload.get("type"),
            }
        )
    return JSONResponse(
        {
            "run_id": run_id,
            "run_status": run_row.get("status"),
            "percent_complete": percent,
            "completed_plugins": completed,
            "total_plugins": total,
            "plugins": plugin_statuses,
            "errors": errors,
        }
    )


@app.get("/api/jobs")
async def list_jobs(status: str = "queued") -> JSONResponse:
    return JSONResponse({"jobs": pipeline.storage.list_analysis_jobs(status=status)})


@app.post("/datasets/{dataset_version_id}/run", response_class=HTMLResponse)
async def run_dataset_plugin(
    request: Request,
    dataset_version_id: str,
    plugins: str = Form(""),
    settings_json: str = Form(""),
    run_seed: int | None = Form(None),
) -> HTMLResponse:
    plugin_ids = [p for p in plugins.split(",") if p]
    settings = {}
    if settings_json:
        try:
            settings = json.loads(settings_json)
        except json.JSONDecodeError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
    if run_seed in (None, 0):
        run_seed = _auto_seed(dataset_version_id, 0)
    run_id = pipeline.run(
        None,
        plugin_ids,
        settings,
        int(run_seed or 0),
        dataset_version_id=dataset_version_id,
    )
    return HTMLResponse(
        f"Run started: {run_id}. <a href='/runs/{run_id}'>View</a>"
    )


@app.post("/datasets/{dataset_version_id}/backfill", response_class=HTMLResponse)
async def backfill_dataset(
    request: Request,
    dataset_version_id: str,
    plugin_id: str = Form(...),
) -> HTMLResponse:
    spec_map = {s.plugin_id: s for s in pipeline.manager.discover()}
    if plugin_id not in spec_map:
        raise HTTPException(status_code=404, detail="Plugin not found")
    spec = spec_map[plugin_id]
    module_path, _ = spec.entrypoint.split(":", 1)
    if module_path.endswith(".py"):
        module_file = spec.path / module_path
    else:
        module_file = spec.path / f"{module_path}.py"
    code_hash = file_sha256(module_file) if module_file.exists() else None
    settings_hash = hashlib.sha256(json_dumps({}).encode("utf-8")).hexdigest()
    pipeline.storage.enqueue_analysis_job(
        dataset_version_id,
        plugin_id,
        spec.version,
        code_hash,
        settings_hash,
        0,
        now_iso(),
    )
    _run_job_queue()
    return HTMLResponse("Backfill queued. <a href='/projects'>Back</a>")


def _run_job_queue() -> None:
    jobs = pipeline.storage.list_analysis_jobs(status="queued")
    for job in jobs:
        job_id = int(job["job_id"])
        pipeline.storage.update_analysis_job_status(
            job_id, "running", started_at=now_iso()
        )
        try:
            pipeline.run(
                None,
                [job["plugin_id"]],
                {},
                int(job.get("run_seed") or 0),
                dataset_version_id=job["dataset_version_id"],
            )
            pipeline.storage.update_analysis_job_status(
                job_id, "completed", completed_at=now_iso()
            )
        except Exception as exc:  # pragma: no cover - failure path
            pipeline.storage.update_analysis_job_status(
                job_id, "error", completed_at=now_iso(), error={"message": str(exc)}
            )


@app.post("/api/jobs/run")
async def run_jobs(background: BackgroundTasks) -> JSONResponse:
    background.add_task(_run_job_queue)
    return JSONResponse({"status": "queued"})


@app.post("/api/backfill")
async def backfill(payload: dict, background: BackgroundTasks) -> JSONResponse:
    plugin_id = payload.get("plugin_id")
    dataset_version_id = payload.get("dataset_version_id")
    if not plugin_id:
        raise HTTPException(status_code=400, detail="plugin_id required")
    spec_map = {s.plugin_id: s for s in pipeline.manager.discover()}
    if plugin_id not in spec_map:
        raise HTTPException(status_code=404, detail="Plugin not found")
    spec = spec_map[plugin_id]
    module_path, _ = spec.entrypoint.split(":", 1)
    if module_path.endswith(".py"):
        module_file = spec.path / module_path
    else:
        module_file = spec.path / f"{module_path}.py"
    code_hash = file_sha256(module_file) if module_file.exists() else None
    settings_hash = hashlib.sha256(json_dumps({}).encode("utf-8")).hexdigest()

    if dataset_version_id:
        pipeline.storage.enqueue_analysis_job(
            dataset_version_id,
            plugin_id,
            spec.version,
            code_hash,
            settings_hash,
            0,
            now_iso(),
        )
    else:
        for dataset in pipeline.storage.list_dataset_versions():
            pipeline.storage.enqueue_analysis_job(
                dataset["dataset_version_id"],
                plugin_id,
                spec.version,
                code_hash,
                settings_hash,
                0,
                now_iso(),
            )
    background.add_task(_run_job_queue)
    return JSONResponse({"status": "queued"})


@app.get("/api/trace")
async def trace(entity_type: str, key: str, max_depth: int = 5) -> JSONResponse:
    return JSONResponse(
        pipeline.storage.trace_from_entity(entity_type, key, max_depth)
    )


@app.get("/trace", response_class=HTMLResponse)
async def trace_view(
    request: Request,
    entity_type: str | None = None,
    key: str | None = None,
    max_depth: int = 5,
) -> HTMLResponse:
    result = ""
    if entity_type and key:
        payload = pipeline.storage.trace_from_entity(entity_type, key, max_depth)
        result = json_dumps(payload)
    return TEMPLATES.TemplateResponse(
        "trace.html",
        {
            "request": request,
            "entity_type": entity_type or "",
            "key": key or "",
            "max_depth": max_depth,
            "result": result,
        },
    )


@app.post("/datasets/{dataset_version_id}/deliver", response_class=HTMLResponse)
async def deliver_dataset(
    request: Request,
    dataset_version_id: str,
    plugin_id: str = Form(""),
    notes: str = Form(""),
) -> HTMLResponse:
    dataset = pipeline.storage.get_dataset_version_context(dataset_version_id)
    if not dataset:
        raise HTTPException(status_code=404, detail="Dataset not found")
    dataset_hash = dataset.get("data_hash")
    specs = {s.plugin_id: s for s in pipeline.manager.discover() if s.type == "analysis"}
    results = pipeline.storage.fetch_latest_plugin_results_for_dataset(
        dataset_version_id
    )
    results_map = {row["plugin_id"]: row for row in results}

    target_plugins = []
    if plugin_id:
        if plugin_id not in specs:
            raise HTTPException(status_code=404, detail="Plugin not found")
        target_plugins = [plugin_id]
    else:
        target_plugins = list(results_map.keys())

    for pid in target_plugins:
        result = results_map.get(pid)
        if not result:
            continue
        spec = specs.get(pid)
        module_path, _ = spec.entrypoint.split(":", 1)
        if module_path.endswith(".py"):
            module_file = spec.path / module_path
        else:
            module_file = spec.path / f"{module_path}.py"
        code_hash = file_sha256(module_file) if module_file.exists() else None
        pipeline.storage.record_delivery(
            dataset["project_id"],
            dataset_version_id,
            pid,
            result.get("plugin_version") or spec.version,
            code_hash,
            dataset_hash,
            now_iso(),
            notes or None,
        )
    return HTMLResponse("Delivery recorded. <a href='/projects'>Back</a>")


@app.post("/api/datasets/{dataset_version_id}/deliveries")
async def mark_delivery(
    dataset_version_id: str,
    plugin_id: str = Form(""),
    notes: str = Form(""),
) -> JSONResponse:
    dataset = pipeline.storage.get_dataset_version_context(dataset_version_id)
    if not dataset:
        raise HTTPException(status_code=404, detail="Dataset not found")
    dataset_hash = dataset.get("data_hash")
    specs = {s.plugin_id: s for s in pipeline.manager.discover() if s.type == "analysis"}
    results = pipeline.storage.fetch_latest_plugin_results_for_dataset(
        dataset_version_id
    )
    results_map = {row["plugin_id"]: row for row in results}

    target_plugins = []
    if plugin_id:
        if plugin_id not in specs:
            raise HTTPException(status_code=404, detail="Plugin not found")
        target_plugins = [plugin_id]
    else:
        target_plugins = list(results_map.keys())

    if not target_plugins:
        raise HTTPException(status_code=400, detail="No plugins to deliver")

    for pid in target_plugins:
        result = results_map.get(pid)
        if not result:
            continue
        spec = specs.get(pid)
        code_hash = _plugin_code_hash(spec) if spec else None
        pipeline.storage.record_delivery(
            dataset["project_id"],
            dataset_version_id,
            pid,
            result.get("plugin_version") or (spec.version if spec else None),
            code_hash,
            dataset_hash,
            now_iso(),
            notes or None,
        )
    return JSONResponse({"status": "ok", "delivered": target_plugins})


@app.get("/runs/{run_id}", response_class=HTMLResponse)
async def run_status(request: Request, run_id: str) -> HTMLResponse:
    run_row = pipeline.storage.fetch_run(run_id)
    project = None
    run_label = ""
    evaluation = _load_evaluation_summary(run_id)
    if run_row and run_row.get("project_id"):
        project = pipeline.storage.fetch_project(run_row["project_id"])
        runs = _annotate_runs(
            pipeline.storage.list_runs_by_project(run_row["project_id"]), project
        )
        for entry in runs:
            if entry.get("run_id") == run_id:
                run_label = str(entry.get("label") or "")
                break
    return TEMPLATES.TemplateResponse(
        "run.html",
        {
            "request": request,
            "run_id": run_id,
            "run_label": run_label,
            "project": project,
            "evaluation": evaluation,
        },
    )


@app.get("/runs/{run_id}/report", response_class=HTMLResponse)
async def run_report_view(request: Request, run_id: str) -> HTMLResponse:
    run_row = pipeline.storage.fetch_run(run_id)
    if not run_row:
        raise HTTPException(status_code=404, detail="Run not found")
    run_dir = pipeline.base_dir / "runs" / run_id
    report_path = run_dir / "report.json"
    if report_path.exists():
        report = json.loads(report_path.read_text(encoding="utf-8"))
    else:
        report = build_report(
            pipeline.storage, run_id, run_dir, Path("docs/report.schema.json")
        )
        write_report(report, run_dir)
    project = None
    if run_row.get("project_id"):
        project = pipeline.storage.fetch_project(run_row["project_id"])
    runs = []
    if project:
        runs = _annotate_runs(pipeline.storage.list_runs_by_project(project["project_id"]), project)
    run_label = ""
    if runs:
        for entry in runs:
            if entry.get("run_id") == run_id:
                run_label = str(entry.get("label") or "")
                break
    evaluation = _load_evaluation_summary(run_id)
    insights = _build_insights(report)
    specs = {spec.plugin_id: spec for spec in pipeline.manager.discover()}
    plugin_labels = {pid: spec.name for pid, spec in specs.items()}
    plugin_descriptions = {
        pid: str(spec.settings.get("description") or "") for pid, spec in specs.items()
    }
    upload_row = None
    if run_row.get("upload_id"):
        upload_row = pipeline.storage.fetch_upload(run_row["upload_id"])
    known_payload = None
    if project and project.get("erp_type"):
        erp_type = str(project.get("erp_type") or "unknown").strip() or "unknown"
        known_payload = _load_known_issues("erp_type", erp_type)
    if known_payload is None and upload_row and upload_row.get("sha256"):
        known_payload = _load_known_issues("sha256", str(upload_row.get("sha256")))
    known_results = _build_known_issue_results(report, known_payload)
    report_errors = []
    for plugin_id, plugin in report.get("plugins", {}).items():
        if plugin.get("error"):
            report_errors.append(
                {
                    "plugin_id": plugin_id,
                    "message": plugin["error"].get("message"),
                    "type": plugin["error"].get("type"),
                }
            )
    return TEMPLATES.TemplateResponse(
        "report.html",
        {
            "request": request,
            "run_id": run_id,
            "report": report,
            "project": project,
            "insights": insights,
            "run_label": run_label,
            "evaluation": evaluation,
            "known_results": known_results,
            "plugin_labels": plugin_labels,
            "plugin_descriptions": plugin_descriptions,
            "report_errors": report_errors,
        },
    )


@app.get("/runs/{run_id}/evaluate", response_class=HTMLResponse)
async def run_evaluate_view(
    request: Request, run_id: str, template: int = 0, known: int = 0
) -> HTMLResponse:
    run_row = pipeline.storage.fetch_run(run_id)
    if not run_row:
        raise HTTPException(status_code=404, detail="Run not found")
    run_dir = pipeline.base_dir / "runs" / run_id
    report_path = run_dir / "report.json"
    if report_path.exists():
        report = json.loads(report_path.read_text(encoding="utf-8"))
    else:
        report = build_report(
            pipeline.storage, run_id, run_dir, Path("docs/report.schema.json")
        )
        write_report(report, run_dir)
    ground_truth = ""
    ground_truth_source = ""
    upload_row = (
        pipeline.storage.fetch_upload(run_row["upload_id"])
        if run_row.get("upload_id")
        else None
    )
    project_row = None
    if run_row.get("dataset_version_id"):
        ctx = pipeline.storage.get_dataset_version_context(run_row["dataset_version_id"])
        if ctx and ctx.get("project_id"):
            project_row = pipeline.storage.fetch_project(ctx["project_id"])
    mode = "template" if template else "known" if known else "auto"
    ground_truth, ground_truth_source = _resolve_ground_truth(
        mode, report, upload_row, project_row, None
    )
    if not ground_truth and (known or template):
        ground_truth = _ground_truth_template(report)
        ground_truth_source = "template"
    last_eval = _load_evaluation_result(run_dir) or {}
    result = last_eval.get("result")
    messages = last_eval.get("messages") if isinstance(last_eval.get("messages"), list) else []
    evaluated_at = last_eval.get("evaluated_at") or ""
    return TEMPLATES.TemplateResponse(
        "evaluate.html",
        {
            "request": request,
            "run_id": run_id,
            "run_status": run_row.get("status") or "",
            "ground_truth": ground_truth,
            "ground_truth_source": ground_truth_source,
            "upload_id": run_row.get("upload_id") or "",
            "result": result,
            "messages": messages,
            "evaluated_at": evaluated_at,
        },
    )


@app.post("/runs/{run_id}/evaluate", response_class=HTMLResponse)
async def run_evaluate(
    request: Request, run_id: str, ground_truth: str = Form(...)
) -> HTMLResponse:
    run_row = pipeline.storage.fetch_run(run_id)
    if not run_row:
        raise HTTPException(status_code=404, detail="Run not found")
    run_dir = pipeline.base_dir / "runs" / run_id
    report_path = run_dir / "report.json"
    if report_path.exists():
        report = json.loads(report_path.read_text(encoding="utf-8"))
    else:
        report = build_report(
            pipeline.storage, run_id, run_dir, Path("docs/report.schema.json")
        )
        write_report(report, run_dir)
    gt_path = run_dir / "ground_truth.yaml"
    gt_path.write_text(ground_truth, encoding="utf-8")
    ok, messages = evaluate_report(report_path, gt_path)
    eval_payload = _write_evaluation(run_dir, ok, messages)
    result = eval_payload["result"]
    return TEMPLATES.TemplateResponse(
        "evaluate.html",
        {
            "request": request,
            "run_id": run_id,
            "run_status": run_row.get("status") or "",
            "ground_truth": ground_truth,
            "ground_truth_source": "manual",
            "upload_id": run_row.get("upload_id") or "",
            "result": result,
            "messages": messages,
            "evaluated_at": eval_payload["evaluated_at"],
        },
    )


@app.post("/api/runs/{run_id}/evaluate")
async def api_run_evaluate(run_id: str, payload: dict) -> JSONResponse:
    run_row = pipeline.storage.fetch_run(run_id)
    if not run_row:
        raise HTTPException(status_code=404, detail="Run not found")
    run_dir = APPDATA_DIR / "runs" / run_id
    report_path = run_dir / "report.json"
    if report_path.exists():
        report = json.loads(report_path.read_text(encoding="utf-8"))
    else:
        report = build_report(
            pipeline.storage, run_id, run_dir, Path("docs/report.schema.json")
        )
        write_report(report, run_dir)

    upload_row = (
        pipeline.storage.fetch_upload(run_row["upload_id"])
        if run_row.get("upload_id")
        else None
    )
    project_row = None
    if run_row.get("dataset_version_id"):
        ctx = pipeline.storage.get_dataset_version_context(run_row["dataset_version_id"])
        if ctx and ctx.get("project_id"):
            project_row = pipeline.storage.fetch_project(ctx["project_id"])
    mode = str(payload.get("mode") or "auto")
    provided = payload.get("ground_truth")
    if provided is not None and not isinstance(provided, str):
        provided = None
    ground_truth, source = _resolve_ground_truth(
        mode, report, upload_row, project_row, provided
    )
    gt_path = run_dir / "ground_truth.yaml"
    gt_path.write_text(ground_truth or "", encoding="utf-8")
    ok, messages = evaluate_report(report_path, gt_path)
    eval_payload = _write_evaluation(run_dir, ok, messages)
    eval_payload["ground_truth_source"] = source
    return JSONResponse(eval_payload)


@app.get("/api/runs/{run_id}/report.json")
async def get_report_json(run_id: str) -> FileResponse:
    report_path = APPDATA_DIR / "runs" / run_id / "report.json"
    if not report_path.exists():
        raise HTTPException(status_code=404, detail="Report not found")
    return FileResponse(report_path)


@app.get("/api/runs/{run_id}/evaluation")
async def get_evaluation(run_id: str) -> JSONResponse:
    eval_path = APPDATA_DIR / "runs" / run_id / "evaluation.json"
    if not eval_path.exists():
        raise HTTPException(status_code=404, detail="Evaluation not found")
    try:
        payload = json.loads(eval_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    if not isinstance(payload, dict):
        raise HTTPException(status_code=400, detail="Invalid evaluation payload")
    return JSONResponse(payload)


@app.get("/api/runs/{run_id}/report.md")
async def get_report_md(run_id: str) -> FileResponse:
    report_path = APPDATA_DIR / "runs" / run_id / "report.md"
    if not report_path.exists():
        raise HTTPException(status_code=404, detail="Report not found")
    return FileResponse(report_path)


@app.get("/api/runs/{run_id}/business_summary.md")
async def get_business_summary(run_id: str) -> FileResponse:
    summary_path = APPDATA_DIR / "runs" / run_id / "business_summary.md"
    if not summary_path.exists():
        raise HTTPException(status_code=404, detail="Business summary not found")
    return FileResponse(summary_path)


@app.get("/api/runs/{run_id}/engineering_summary.md")
async def get_engineering_summary(run_id: str) -> FileResponse:
    summary_path = APPDATA_DIR / "runs" / run_id / "engineering_summary.md"
    if not summary_path.exists():
        raise HTTPException(status_code=404, detail="Engineering summary not found")
    return FileResponse(summary_path)


@app.get("/api/runs/{run_id}/appendix_raw.md")
async def get_appendix_raw(run_id: str) -> FileResponse:
    summary_path = APPDATA_DIR / "runs" / run_id / "appendix_raw.md"
    if not summary_path.exists():
        raise HTTPException(status_code=404, detail="Appendix not found")
    return FileResponse(summary_path)


@app.get("/api/runs/{run_id}/slide_kit/{artifact_path:path}")
async def get_slide_kit(run_id: str, artifact_path: str) -> FileResponse:
    base = APPDATA_DIR / "runs" / run_id / "slide_kit"
    try:
        resolved = safe_join(base, artifact_path)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    if not resolved.exists():
        raise HTTPException(status_code=404, detail="Slide kit file not found")
    return FileResponse(resolved)


@app.get("/api/runs/{run_id}")
async def get_run_status(run_id: str) -> JSONResponse:
    run_row = pipeline.storage.fetch_run(run_id)
    if not run_row:
        raise HTTPException(status_code=404, detail="Run not found")
    return JSONResponse({"run": run_row})


@app.get("/api/runs/{run_id}/artifacts/{plugin_id}/{artifact_path:path}")
async def get_artifact(run_id: str, plugin_id: str, artifact_path: str) -> FileResponse:
    base = APPDATA_DIR / "runs" / run_id / "artifacts" / plugin_id
    try:
        resolved = safe_join(base, artifact_path)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    if not resolved.exists():
        raise HTTPException(status_code=404, detail="Artifact not found")
    return FileResponse(resolved)

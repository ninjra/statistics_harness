from __future__ import annotations

import hashlib
import json
from pathlib import Path
from typing import Any, Iterable

import pandas as pd

from statistic_harness.core.types import PluginArtifact, PluginResult
from statistic_harness.core.utils import json_dumps, now_iso, write_json


class Plugin:
    def run(self, ctx) -> PluginResult:
        input_file = ctx.settings.get("input_file")
        if not input_file:
            raise ValueError("Input file not found")
        path = Path(input_file)
        if not path.exists():
            raise ValueError("Input file not found")

        encoding = ctx.settings.get("encoding", "utf-8")
        delimiter = ctx.settings.get("delimiter")
        sheet_name = ctx.settings.get("sheet_name")
        # Larger default chunk improves ingest throughput for multi-million row datasets.
        chunk_size = int(ctx.settings.get("chunk_size", 10_000))

        dataset_version_id = ctx.dataset_version_id
        if not dataset_version_id:
            raise ValueError("Dataset version not provided")
        existing = ctx.storage.get_dataset_version(dataset_version_id)
        if existing:
            existing_hash = existing.get("data_hash")
            same_media = False
            if ctx.input_hash and existing_hash and existing_hash == ctx.input_hash:
                same_media = True
            elif (
                not existing_hash
                and ctx.dataset_id
                and existing.get("dataset_id") == ctx.dataset_id
            ):
                same_media = True
            if same_media:
                ctx.storage.reset_dataset_version(
                    dataset_version_id,
                    table_name=existing.get("table_name")
                    or f"dataset_{dataset_version_id}",
                    data_hash=ctx.input_hash,
                    created_at=now_iso(),
                )
                existing = ctx.storage.get_dataset_version(dataset_version_id)
        if existing and int(existing.get("row_count") or 0) > 0:
            columns = ctx.storage.fetch_dataset_columns(dataset_version_id)
            schema = {
                "columns": [
                    {
                        "name": col["original_name"],
                        "dtype": col.get("dtype"),
                    }
                    for col in columns
                ]
            }
            schema_path = ctx.run_dir / "dataset" / "schema.json"
            write_json(schema_path, schema)
            return PluginResult(
                status="ok",
                summary=f"Dataset already ingested ({existing['row_count']} rows)",
                metrics={
                    "rows": int(existing.get("row_count") or 0),
                    "cols": int(existing.get("column_count") or 0),
                },
                findings=[],
                artifacts=[
                    PluginArtifact(
                        path=str(schema_path.relative_to(ctx.run_dir)),
                        type="json",
                        description="Schema",
                    )
                ],
                error=None,
            )

        def dtype_to_sqlite(dtype: Any) -> str:
            if pd.api.types.is_integer_dtype(dtype):
                return "INTEGER"
            if pd.api.types.is_float_dtype(dtype):
                return "REAL"
            if pd.api.types.is_bool_dtype(dtype):
                return "INTEGER"
            return "TEXT"

        def normalize_value(value: Any) -> Any:
            if pd.isna(value):
                return None
            if isinstance(value, pd.Timestamp):
                return value.isoformat()
            return value

        def build_columns(columns: list[str], dtypes: list[Any]) -> list[dict[str, Any]]:
            # Disambiguate duplicate headers so the DB metadata is stable and lossless.
            safe_names = [f"c{idx+1}" for idx in range(len(columns))]
            out: list[dict[str, Any]] = []
            name_counts: dict[str, int] = {}
            for idx, (orig, safe, dtype) in enumerate(zip(columns, safe_names, dtypes), start=1):
                source = str(orig) if orig is not None else ""
                base = source.strip() or f"column_{idx}"
                if base in name_counts:
                    name_counts[base] += 1
                    internal = f"{base}_{name_counts[base]}"
                else:
                    name_counts[base] = 1
                    internal = base
                out.append(
                    {
                        "column_id": idx,
                        "safe_name": safe,
                        # Internal unique name used throughout the harness.
                        "original_name": internal,
                        # Preserve the raw header for traceability.
                        "source_original_name": source,
                        "dtype": str(dtype),
                        "sqlite_type": dtype_to_sqlite(dtype),
                    }
                )
            return out

        def iter_csv_chunks() -> Iterable[pd.DataFrame]:
            return pd.read_csv(
                path, delimiter=delimiter, encoding=encoding, chunksize=chunk_size
            )

        def iter_json_chunks() -> Iterable[pd.DataFrame]:
            try:
                for chunk in pd.read_json(path, lines=True, chunksize=chunk_size):
                    yield chunk
                return
            except ValueError:
                df = pd.read_json(path)
                yield df

        def iter_excel_rows() -> Iterable[list[Any]]:
            import mimetypes

            # Prevent mimetypes from probing system files blocked by the sandbox.
            mimetypes.knownfiles = []
            from openpyxl import load_workbook

            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb[sheet_name] if sheet_name else wb.active
            rows = ws.iter_rows(values_only=True)
            try:
                headers = next(rows)
            except StopIteration:
                return
            headers = [
                (str(h) if h is not None else f"column_{idx+1}")
                for idx, h in enumerate(headers)
            ]
            yield headers
            for row in rows:
                yield list(row)

        columns_meta: list[dict[str, Any]] = []
        row_index = 0
        table_name = f"dataset_{dataset_version_id}"
        # Full-dataset abstract representation (computed during ingest to avoid re-scans):
        # - null counts for every column
        # - numeric min/max/mean (via running sum)
        # - text min/max length
        # - text numeric-like ratio (for later normalization decisions)
        stats_by_safe: dict[str, dict[str, Any]] = {}
        internal_to_safe: dict[str, str] = {}

        def _ensure_stats_initialized() -> None:
            nonlocal stats_by_safe, internal_to_safe
            if stats_by_safe or not columns_meta:
                return
            for col in columns_meta:
                internal = str(col.get("original_name") or "")
                safe = str(col.get("safe_name") or "")
                sqlite_type = str(col.get("sqlite_type") or "TEXT").upper()
                if not internal or not safe:
                    continue
                internal_to_safe[internal] = safe
                stats_by_safe[safe] = {
                    "sqlite_type": sqlite_type,
                    "original_name": internal,
                    "source_original_name": str(col.get("source_original_name") or internal),
                    "n": 0,
                    "nulls": 0,
                }
                if sqlite_type in {"INTEGER", "REAL"}:
                    stats_by_safe[safe].update(
                        {
                            "min": None,
                            "max": None,
                            "sum": 0.0,
                        }
                    )
                else:
                    stats_by_safe[safe].update(
                        {
                            "min_len": None,
                            "max_len": None,
                            "numeric_like_total": 0,
                            "numeric_like_num": 0,
                        }
                    )

        def _update_stats_from_frame(df: pd.DataFrame) -> None:
            _ensure_stats_initialized()
            if df is None or df.empty:
                return
            for internal, safe in internal_to_safe.items():
                if internal not in df.columns:
                    continue
                st = stats_by_safe[safe]
                s = df[internal]
                n = int(len(s))
                st["n"] = int(st.get("n") or 0) + n
                nulls = int(s.isna().sum())
                st["nulls"] = int(st.get("nulls") or 0) + nulls
                sqlite_type = str(st.get("sqlite_type") or "TEXT").upper()
                if sqlite_type in {"INTEGER", "REAL"}:
                    # Use pandas numeric conversion for robustness; errors become NaN (counted as nulls).
                    sn = pd.to_numeric(s, errors="coerce")
                    # Min/max across non-null values.
                    if sn.notna().any():
                        cur_min = float(sn.min())
                        cur_max = float(sn.max())
                        prev_min = st.get("min")
                        prev_max = st.get("max")
                        st["min"] = cur_min if prev_min is None else float(min(prev_min, cur_min))
                        st["max"] = cur_max if prev_max is None else float(max(prev_max, cur_max))
                        st["sum"] = float(st.get("sum") or 0.0) + float(sn.sum())
                else:
                    ss = s.dropna().astype(str)
                    if not ss.empty:
                        lens = ss.str.len()
                        cur_min = int(lens.min())
                        cur_max = int(lens.max())
                        prev_min = st.get("min_len")
                        prev_max = st.get("max_len")
                        st["min_len"] = cur_min if prev_min is None else int(min(prev_min, cur_min))
                        st["max_len"] = cur_max if prev_max is None else int(max(prev_max, cur_max))
                        # Numeric-like detection for TEXT columns (matches transform_normalize_mixed intent).
                        # This scans all values during ingest (no sampling) and enables later DB normalization
                        # to avoid a second full-table scan.
                        cleaned = ss.str.strip()
                        cleaned = cleaned.str.replace(",", "", regex=False).str.replace("_", "", regex=False)
                        try:
                            matches = cleaned.str.match(r"^[+-]?(\d+(\.\d+)?|\.\d+)$", na=False)
                            st["numeric_like_total"] = int(st.get("numeric_like_total") or 0) + int(len(cleaned))
                            st["numeric_like_num"] = int(st.get("numeric_like_num") or 0) + int(matches.sum())
                        except Exception:
                            pass

        with ctx.storage.connection() as conn:
            if path.suffix.lower() in {".xlsx"}:
                row_iter = iter_excel_rows()
                headers = next(row_iter, None)
                if headers is None:
                    raise ValueError("No rows found")
                raw_headers = list(headers)
                sample = []
                for _ in range(chunk_size):
                    row = next(row_iter, None)
                    if row is None:
                        break
                    sample.append(row)
                sample_df = pd.DataFrame(sample, columns=raw_headers)
                columns_meta = build_columns(
                    list(sample_df.columns), list(sample_df.dtypes)
                )
                internal_headers = [c["original_name"] for c in columns_meta]
                sample_df = sample_df.copy()
                sample_df.columns = internal_headers
                _update_stats_from_frame(sample_df)
                ctx.storage.create_dataset_table(table_name, columns_meta, conn)
                ctx.storage.add_append_only_triggers(table_name, conn)
                ctx.storage.replace_dataset_columns(dataset_version_id, columns_meta, conn)

                def emit_rows(rows: list[list[Any]]) -> None:
                    nonlocal row_index
                    safe_columns = [col["safe_name"] for col in columns_meta]
                    batch = []
                    for row in rows:
                        values = [normalize_value(v) for v in row]
                        row_dict = dict(zip(internal_headers, values))
                        row_json = json.dumps(row_dict, ensure_ascii=False)
                        batch.append((row_index, row_json, *values))
                        row_index += 1
                    ctx.storage.insert_dataset_rows(table_name, safe_columns, batch, conn)

                emit_rows(sample)
                buffer = []
                for row in row_iter:
                    buffer.append(row)
                    if len(buffer) >= chunk_size:
                        _update_stats_from_frame(pd.DataFrame(buffer, columns=internal_headers))
                        emit_rows(buffer)
                        buffer = []
                if buffer:
                    _update_stats_from_frame(pd.DataFrame(buffer, columns=internal_headers))
                    emit_rows(buffer)
                ctx.storage.update_dataset_version_stats(
                    dataset_version_id,
                    row_index,
                    len(columns_meta),
                    conn,
                )
            elif path.suffix.lower() == ".json":
                internal_columns: list[str] | None = None
                for chunk in iter_json_chunks():
                    if not columns_meta:
                        columns_meta = build_columns(
                            list(chunk.columns), list(chunk.dtypes)
                        )
                        internal_columns = [c["original_name"] for c in columns_meta]
                        ctx.storage.create_dataset_table(table_name, columns_meta, conn)
                        ctx.storage.add_append_only_triggers(table_name, conn)
                        ctx.storage.replace_dataset_columns(
                            dataset_version_id, columns_meta, conn
                        )
                    if internal_columns is None:
                        internal_columns = [c["original_name"] for c in columns_meta]
                    chunk = chunk.copy()
                    chunk.columns = internal_columns
                    _update_stats_from_frame(chunk)
                    safe_columns = [col["safe_name"] for col in columns_meta]
                    batch = []
                    for row in chunk.itertuples(index=False, name=None):
                        values = [normalize_value(v) for v in row]
                        row_dict = dict(zip(chunk.columns, values))
                        row_json = json.dumps(row_dict, ensure_ascii=False)
                        batch.append((row_index, row_json, *values))
                        row_index += 1
                    ctx.storage.insert_dataset_rows(table_name, safe_columns, batch, conn)
                ctx.storage.update_dataset_version_stats(
                    dataset_version_id, row_index, len(columns_meta), conn
                )
            else:
                internal_columns: list[str] | None = None
                for chunk in iter_csv_chunks():
                    if not columns_meta:
                        columns_meta = build_columns(
                            list(chunk.columns), list(chunk.dtypes)
                        )
                        internal_columns = [c["original_name"] for c in columns_meta]
                        ctx.storage.create_dataset_table(table_name, columns_meta, conn)
                        ctx.storage.add_append_only_triggers(table_name, conn)
                        ctx.storage.replace_dataset_columns(
                            dataset_version_id, columns_meta, conn
                        )
                    if internal_columns is None:
                        internal_columns = [c["original_name"] for c in columns_meta]
                    chunk = chunk.copy()
                    chunk.columns = internal_columns
                    _update_stats_from_frame(chunk)
                    safe_columns = [col["safe_name"] for col in columns_meta]
                    batch = []
                    for row in chunk.itertuples(index=False, name=None):
                        values = [normalize_value(v) for v in row]
                        row_dict = dict(zip(chunk.columns, values))
                        row_json = json.dumps(row_dict, ensure_ascii=False)
                        batch.append((row_index, row_json, *values))
                        row_index += 1
                    ctx.storage.insert_dataset_rows(table_name, safe_columns, batch, conn)
                ctx.storage.update_dataset_version_stats(
                    dataset_version_id, row_index, len(columns_meta), conn
                )

            # Post-ingest DB normalization (full-dataset, DB-backed):
            # - Ensure performance-critical row_index index exists.
            # - ANALYZE to improve query planning for subsequent plugin scans.
            # - Compute per-column stats once on the full dataset and store in metadata.
            try:
                ctx.storage.ensure_dataset_row_index_index(table_name, conn)
            except Exception:
                pass
            try:
                ctx.storage.analyze_table(table_name, conn)
            except Exception:
                pass
            try:
                # Finalize derived stats (mean) from running totals.
                finalized: dict[str, dict[str, Any]] = {}
                for safe, st in stats_by_safe.items():
                    out = dict(st)
                    sqlite_type = str(out.get("sqlite_type") or "TEXT").upper()
                    if sqlite_type in {"INTEGER", "REAL"}:
                        n = int(out.get("n") or 0)
                        nulls = int(out.get("nulls") or 0)
                        used = max(0, n - nulls)
                        if used > 0:
                            out["mean"] = float(out.get("sum") or 0.0) / float(used)
                        out.pop("sum", None)
                    else:
                        total = int(out.get("numeric_like_total") or 0)
                        num = int(out.get("numeric_like_num") or 0)
                        if total > 0:
                            out["numeric_like_ratio"] = float(num) / float(total)
                        out.pop("numeric_like_total", None)
                        out.pop("numeric_like_num", None)
                    finalized[safe] = out

                # Best-effort role hints + indexes to speed downstream plugin SQL filtering/grouping.
                role_by_safe: dict[str, str] = {}

                def infer_role(col_name: str, sqlite_type: str) -> str | None:
                    lname = col_name.lower()
                    if any(tok in lname for tok in ("param", "parameter", "params", "meta", "config")):
                        return "parameter"
                    if any(tok in lname for tok in ("timestamp", "time", "date")):
                        return "timestamp"
                    if lname.endswith("id") or "_id" in lname or " id" in lname:
                        return "id"
                    if any(tok in lname for tok in ("event", "action", "activity")):
                        return "event"
                    if "variant" in lname:
                        return "variant"
                    if "status" in lname:
                        return "status"
                    if sqlite_type in {"INTEGER", "REAL"}:
                        return "numeric"
                    return None

                index_candidates: list[tuple[str, str]] = []  # (orig_name, safe_name)
                for col in columns_meta:
                    orig = str(col.get("original_name") or "")
                    safe = str(col.get("safe_name") or "")
                    sqlite_type = str(col.get("sqlite_type") or "TEXT").upper()
                    if not orig or not safe:
                        continue
                    role = infer_role(orig, sqlite_type)
                    if role:
                        role_by_safe[safe] = role
                    st = finalized.get(safe) or {}
                    max_len = st.get("max_len")
                    # Index short-ish text columns and key-like columns; avoid indexing huge blobs.
                    if role in {"timestamp", "id", "event", "variant", "status"}:
                        if sqlite_type in {"INTEGER", "REAL"}:
                            index_candidates.append((orig, safe))
                        else:
                            try:
                                if max_len is None or int(max_len) <= 256:
                                    index_candidates.append((orig, safe))
                            except (TypeError, ValueError):
                                index_candidates.append((orig, safe))

                # Add exact categorical summaries for a small set of key-like columns.
                # This is full-dataset and DB-backed, but limited to a few columns to
                # keep normalization predictable.
                from statistic_harness.core.utils import quote_identifier

                safe_table_q = quote_identifier(table_name)

                def _sensitive_name(col_name: str) -> bool:
                    lname = col_name.lower()
                    return any(tok in lname for tok in ("email", "ssn", "phone", "address"))

                for orig, safe in index_candidates[:8]:
                    role = role_by_safe.get(safe)
                    if role not in {"event", "variant", "status"}:
                        continue
                    st = finalized.get(safe) or {}
                    sqlite_type = str(st.get("sqlite_type") or "TEXT").upper()
                    if sqlite_type != "TEXT":
                        continue
                    try:
                        max_len = st.get("max_len")
                        if max_len is not None and int(max_len) > 256:
                            continue
                    except (TypeError, ValueError):
                        pass

                    qcol = quote_identifier(safe)
                    try:
                        row = conn.execute(
                            f"SELECT COUNT(DISTINCT {qcol}) AS d FROM {safe_table_q}"
                        ).fetchone()
                        if row is not None:
                            st["distinct_count"] = int(row["d"] or 0)
                    except Exception:
                        pass

                    if not _sensitive_name(orig):
                        try:
                            rows = conn.execute(
                                f"""
                                SELECT {qcol} AS v, COUNT(*) AS c
                                FROM {safe_table_q}
                                GROUP BY {qcol}
                                ORDER BY c DESC
                                LIMIT 20
                                """
                            ).fetchall()
                            top = []
                            for r in rows or []:
                                v = r["v"]
                                if v is None:
                                    sval = "(null)"
                                else:
                                    sval = str(v)
                                    if len(sval) > 200:
                                        sval = sval[:200] + "..."
                                top.append({"value": sval, "count": int(r["c"] or 0)})
                            if top:
                                st["top_values"] = top
                        except Exception:
                            pass
                    finalized[safe] = st

                ctx.storage.update_dataset_column_stats(dataset_version_id, finalized, conn)

                if role_by_safe:
                    try:
                        ctx.storage.update_dataset_column_roles(dataset_version_id, role_by_safe, conn)
                    except Exception:
                        pass

                # Cap index count to keep normalization predictable.
                index_candidates = index_candidates[:12]
                for _, safe in index_candidates:
                    try:
                        ctx.storage.ensure_dataset_column_index(table_name, safe, conn)
                    except Exception:
                        pass
            except Exception:
                # Stats are best-effort; do not fail ingestion.
                pass

        schema = {
            "columns": [
                {
                    "name": col["original_name"],
                    "dtype": col.get("dtype"),
                }
                for col in columns_meta
            ]
        }
        schema_path = ctx.run_dir / "dataset" / "schema.json"
        write_json(schema_path, schema)

        if ctx.dataset_version_id:
            fingerprint_payload = [
                {
                    "name": col["original_name"].lower().strip(),
                    "dtype": col.get("dtype"),
                }
                for col in columns_meta
            ]
            fingerprint_payload = sorted(
                fingerprint_payload, key=lambda item: item["name"]
            )
            fingerprint = hashlib.sha256(
                json_dumps(fingerprint_payload).encode("utf-8")
            ).hexdigest()
            format_id = ctx.storage.ensure_raw_format(
                fingerprint, name=None, created_at=now_iso()
            )
            ctx.storage.set_dataset_raw_format(ctx.dataset_version_id, format_id)

        return PluginResult(
            status="ok",
            summary=f"Ingested {row_index} rows",
            metrics={"rows": int(row_index), "cols": int(len(columns_meta))},
            findings=[],
            artifacts=[
                PluginArtifact(
                    path=str(schema_path.relative_to(ctx.run_dir)),
                    type="json",
                    description="Schema",
                )
            ],
            error=None,
        )
